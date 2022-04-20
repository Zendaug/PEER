# -*- coding: utf-8 -*-
"""
Created on Fri Mar 11 14:08:46 2022

@author: tbednall
"""

import os
import sys
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import csv
import json
from tkinter import messagebox
import re

def open_url(url):
    '''A function to open a URL that is platform independent.'''
    if sys.platform=='win32':
        os.startfile(url)
    elif sys.platform=='darwin':
        subprocess.Popen(['open', url])
    else:
        try:
            subprocess.Popen(['xdg-open', url])
        except OSError:
            print('Please open a browser on: '+url)
            
def open_file(filename):
    '''A function to open a file/folder that is platform indepedent.'''
    if sys.platform == "win32":
        os.startfile(filename)
    else:
        opener ="open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener, filename])

def clean_text(text):
    if type(text) is not str: return(text)
    text = text.strip()
    return(text.encode('unicode_escape').decode('utf-8'))

def remove_blanks(elements):
    return([x for x in elements if (x != "" and x is not None)])
    
def number_beginning(text):
    number = ""
    for char in text:
        if char.isnumeric():
            number = number + char
        else:
            return(int(number))
    if len(number) > 0: return(int(number))

def find_all(text, search_for):
    '''Returns the position in a string of all incidences of a character'''
    find_list = []
    start = 0
    for a in range(0, text.count(search_for)):
        find_list.append(text.find(search_for, start))
        start = find_list[-1]+1
    return(find_list)

def only_numbers(char, event_type = ""):
    return char.isdigit()

def only_posfloat(char, event_type = ""):
    if char == "." or char.isdigit():
        return True
    else:
        return False

def unique(list1):
    unique_list = []
    for item in list1:
        if item not in unique_list: unique_list.append(item)
    return(unique_list)

def replace_multiple(str_text, rep_list, str_rep = ""):
    for item in rep_list: str_text = str_text.replace(str(item), str_rep)
    return(str_text)

def save_config(config):
    try:
        f = open("config.txt", "w")
        f.write(json.dumps(config, sort_keys=True, separators=(',\n', ':')))
        f.close()  
    except:
        messagebox.showinfo("Error", f"Unable to save configuration file (config.txt). Please check that the PEER folder ({os.getcwd()}) has file write access permissions.")

def first_numeric(list1):
    '''Indicates the first instance of a numeric variable in a list'''
    for n, item in enumerate(list1):
        if type(item) is float: item = int(item)
        if str(item).isdigit(): return(n)
    return(None)

def cleanupURL(URL):
    URL = URL.strip().lower()
    if URL[-1] != "/": URL += "/"
    URL = URL.replace('http:', 'https:')
    return(URL)

def writeCSV(data, filename):
    txt = []
    for x, field in enumerate(data.keys()):
        if x == 0:
            n_cases = len(data[field])
        else:
            txt.append(",")
            if len(data[field]) != n_cases:
                messagebox.showinfo("Error", "Cannot write CSV file. The number of rows appears to be different.")
                sys.exit()
        txt.append("\"{}\"".format(field))
    txt.append("\n")
    for y in range(n_cases):
        for x, field in enumerate(data.keys()):
            if x > 0: txt.append(",")
            if "{}".format(data[field][y]) == "nan":
                txt.append("")
            elif type(data[field][y]) in [int, float]:
                txt.append("{}".format(data[field][y]))
            elif type(data[field][y]) == str:
                txt.append("\"{}\"".format(data[field][y]))
            else:
                txt.append("")
        txt.append("\n")
    f = open(filename, "w")
    f.write("".join(txt[0:-1]))
    f.close()

def writeXL(filename, datasets = [], sheetnames = []):
    wb = Workbook()
    for z in range(len(datasets)):
        if z==0:
            ws = wb.active
            ws.title = sheetnames[0]
        else:
            ws = wb.create_sheet(sheetnames[z])
        for x, field in enumerate(datasets[z].keys(),1):
            ws.cell(row = 1, column = x).value = clean_text(field)
            for y, temp_val in enumerate(datasets[z][field],2):
                ws.cell(row = y, column = x).value = clean_text(temp_val)
        ws.auto_filter.ref = 'A1:' + get_column_letter(ws.max_column) + str(ws.max_row)
        ws.freeze_panes = ws["A2"]
    wb.save(filename)

def cleanupNumber(txt):
    if type(txt) == str:
        txt = txt.strip()
        try:
            return int(txt)
        except ValueError:
            try:
                if int(float(txt)) == float(txt): return(int(float(txt)))
                return(float(txt))
            except:
                return(txt)
    else:
        return(txt)

def readTable(filename, delimiter = ","):
    # Open the file using UTF-8 if it's a CSV file, otherwise UTF-16 if TSV or txt
    enc = "utf_8"
    if delimiter == "\t": enc = "utf-16"
    f = open(filename, encoding = enc)
    readFile = csv.reader((x.replace('\0','').replace('\r',"") for x in f), delimiter = delimiter, quotechar = '"')
    df = {}
    for y, row in enumerate(readFile,1):
        # Check for blank rows
        if len(row) > 0:
            if y == 1:
                for field in row:
                    df[field] = []
            else:
                # Check for blank rows and delete
                non_blank = False
                for x, field in enumerate(df.keys()):
                    if len(row[x]) > 0:
                        non_blank = True
                        break
                if non_blank:
                    for x, field in enumerate(df.keys()):
                        df[field].append(cleanupNumber(row[x]))

    f.close()
    return df

def readXL(filename):
    df = {}
    wb = load_workbook(filename)
    sheet_obj = wb.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    for x in range(1, max_col+1):
        col_name = sheet_obj.cell(row = 1, column = x).value
        df[col_name] = []
        for y in range(2, max_row+1):
            df[col_name].append(sheet_obj.cell(row = y, column = x).value)
    return(df)


def df_byRow(df, id_col):
    '''This function converts the data frame into a records format, such that it is organised via an id number'''
    df2 = {}
    # Check for duplicates
    dups = set([a for a in df[id_col] if df[id_col].count(a) > 1])
    if len(dups) > 0:
        messagebox.showinfo("Error", "Duplicates detected: {}. Please check these IDs and try again.".format(dups))
        return None
                
    for key in df[id_col]:
        df2[cleanupNumber(key)] = {} # Converts to a number if possible
    for x, col in enumerate(df):
        if col != id_col:
            for y, row in enumerate(df[col]):
                df2[df[id_col][y]][col] = cleanupNumber(df[col][y])
    return df2

def collapseList(df, list_var, vars, begin = 1, delete_blank = True, unlist = False):
    '''This function collapses a set of dictionary elements into a list'''
    if type(vars) == list:
        for case in df:
            df[case][list_var] = []
            for var in vars:
                if var in df[case]:
                    if df[case][var] != "" and type(df[case][var]) != type(None):
                        df[case][list_var].append(cleanupNumber(df[case].pop(var))) # Convert to a number if possible; otherwise, use string.
                    else:
                        if delete_blank == False: df[case][list_var].append("")
                        del df[case][var]
            if unlist == True and len(df[case][list_var]) == 1:
                df[case][list_var] = df[case][list_var][0]            
    elif type(vars) == str:
        if "*" in vars:
            for case in df:
                df[case][list_var] = []
                a = begin
                var = vars.replace("*", str(a))
                while var in df[case]:
                    if df[case][var] != "" and type(df[case][var]) != type(None):
                        df[case][list_var].append(cleanupNumber(df[case].pop(var)))
                    else:
                        if delete_blank == False: df[case][list_var].append("")
                        del df[case][var]
                    a = a + 1
                    var = vars.replace("*", str(a))
                if unlist == True and len(df[case][list_var]) == 1:
                    df[case][list_var] = df[case][list_var][0]    
    return(df)

def df_byCol(df, id_col = "ExternalDataReference"):
    df2 = {id_col: []}
    for case in df:
        for col_name in df[case].keys():
            if type(df[case][col_name]) is list: # Parse a list as different columns
                for x in range(len(df[case][col_name])):
                    col_name2 = "{}_{}".format(col_name, x+1)
                    if col_name2 not in df2: df2[col_name2] = []
            else:
                if col_name not in df2: df2[col_name] = []
        
    for case in df.keys():
        for col_name in df2.keys(): df2[col_name].append(None) # Add blank values by default
        df2[id_col].append(case) # Add case number
        for col_name in df[case].keys():
            if type(df[case][col_name]) is list:
                for x in range(len(df[case][col_name])):
                    df2["{}_{}".format(col_name,x+1)][-1] = df[case][col_name][x]
            else:
                df2[col_name][-1] = df[case][col_name]
    return(df2)

def flatten_list(lst):
    lst2 = []
    for item in lst:
        if type(item) is list:
            lst2 = lst2 + flatten_list(item)
        elif item is None or item is str:
            pass
        else:
            lst2 = lst2 + [item]
    return(lst2)

def nansum(lst):
    return(sum(flatten_list(lst)))
    
def nanmean(lst):
    lst2 = flatten_list(lst)
    return(sum(lst2)/len(lst2))

def npmin(lst):
    lst2 = flatten_list(lst)
    return(min(lst2))