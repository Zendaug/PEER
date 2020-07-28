# -*- coding: utf-8 -*-

try:
    f = open("Version History.txt", "r")
    updates = f.read()
    f.close()
    print(updates)
except:
    print("Unable to locate Version History.")

print("\n(C) Tim Bednall 2019-2020.")

import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from tkcalendar import DateEntry
from canvasapi import Canvas
from random import shuffle
import json
from datetime import datetime, timedelta
import os
import requests
import sys
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import csv

#%% Global functions and variables   

canvas = None
qualtrics = ""
orig_data = ""
tw_list = [] # Questions that make up the teamwork score
tw_names = [] # The names of the teamwork variables 
other_list = [] # The names of other variables

# Set up dictionary for ratios and adjusted scores for fast lookup
ratio_dict = {} 
adj_dict = {}

def open_url(url):
    if sys.platform=='win32':
        os.startfile(url)
    elif sys.platform=='darwin':
        subprocess.Popen(['open', url])
    else:
        try:
            subprocess.Popen(['xdg-open', url])
        except OSError:
            print('Please open a browser on: '+url)

def clean_text(text):
    if type(text) is not str: return(text)
    text = text.strip()
    return(text.encode('unicode_escape').decode('utf-8'))

def remove_blanks(elements):
    return([x for x in elements if (x != "" and x is not None)])
    
def number_beginning(text):
    number = ""
    for char in text:
        if char.isnumeric():
            number = number + char
        else:
            return(int(number))
    if len(number) > 0: return(int(number))

def find_all(text, search_for):
    '''Returns the position in a string of all incidences of a character'''
    find_list = []
    start = 0
    for a in range(0, text.count(search_for)):
        find_list.append(text.find(search_for, start))
        start = find_list[-1]+1
    return(find_list)

def only_numbers(char, event_type = ""):
    return char.isdigit()

def unique(list1):
    unique_list = []
    for item in list1:
        if item not in unique_list: unique_list.append(item)
    return(unique_list)

def replace_multiple(str_text, rep_list, str_rep = ""):
    for item in rep_list: str_text = str_text.replace(str(item), str_rep)
    return(str_text)

def save_config():
    try:
        f = open("config.txt", "w")
        f.write(json.dumps(config, sort_keys=True, separators=(',\n', ':')))
        f.close()   
    except:
        print("Unabled to save configuration file (config.txt).")

# Load the configuration file
# First of all, load the hard defaults
hard_defaults = {
    'API_URL': "",
    'API_TOKEN': "",
    'EmailSuffix': "",
    'ScoreType': 2,
    'SelfVote': 2,
    'PeerMark_Name': "Peer Mark",
    'MinimumPeers': 2,
    'Penalty_NonComplete': 100,
    'Penalty_PartialComplete': 0,
    'Penalty_SelfPerfect': 0,
    'Penalty_PeersAllZero': 0,
    'Penalty_PerDayLate': 0,
    'Exclude_PartialComplete': 0,
    'Exclude_SelfPerfect': 0,
    'Exclude_PeersAllZero': 0,
    'SaveLong': 0,
    'SaveWide': 0,
    'PointsPossible': 5,
    'RescaleTo': 5,
    'publish_assignment': True,
    'weekdays_only': True,
    'subject_confirm': "Confirming Group Membership",
    'message_confirm': "Dear [firstnames_list],\n\nI am writing to confirm your membership in: [name].\n\nAccording to our records, the members of this group include:\n\n[members_bulletlist]\n\nIf there are any errors in our records -- such as missing group members or people incorrectly listed as group members -- please contact me as soon as possible. You may ignore this message if these records are correct.\n\nWe encourage you to visit your group's homepage, a platform where you can collaborate with your fellow members. Your group's homepage is located at: [group_homepage]",
    'subject_nogroup' : "Lack of Group Membership",
    'message_nogroup' : "Dear [first_name],\n\nI am writing because according to our records, you are not currently listed as a member of any student team for this unit.\n\nAs you would be aware, there is a group assignment due later in the semester. If you do not belong to a team, you will be unable to complete the assignment.\n\nIf there is another circumstance that would prevent you from joining a group project, please let me know as soon as possible.\n\nPlease let me know if there is anything I can do to support you in the meantime.",
    'subject_invitation': "Peer Evaluation now available",
    'message_invitation': "Dear [First Name],\n\nThe Peer Evaluation survey is now available.\n\nYou may access it via this link: [Link]\n\nThis link is unique to yourself, so please do not share it with other people.",
    'subject_reminder': "Reminder: Peer Evaluation due",
    'message_reminder': "Dear [First Name],\n\nThis is a courtesy reminder to complete the Peer Evaluation survey. According to our records, you have not yet completed this survey.\n\nYou may access the survey via this link: [Link]\n\nThis link is unique to yourself, so please do not share it with other people.",
    'subject_thankyou' : "Thank you for completing the Peer Evaluation",
    'message_thankyou' : "Dear [First Name],\n\nThank you for completing the Peer Evaluation. There is no need to take any further action.",
    'firsttime_export' : True,
    'firsttime_upload' : True,
    'ExportData': True,
    'UploadData': True
    }
config = hard_defaults

# Overwrite hard defaults with soft defaults
try:
    f = open("defaults.txt", "r")
    defaults = json.loads(f.read())
    f.close()
    for item in defaults: config[item] = defaults[item]
except:
    defaults = hard_defaults

# Overwrite defaults with user's own configuration file
try:
    f = open("config.txt", "r")
    user_config = json.loads(f.read())
    f.close()
    for item in user_config: config[item] = user_config[item]
except:
    print("Creating user configuration file from defaults...")

# Overwrite the user's configuration with any updates
try:
    f = open("update.txt", "r")
    user_config = json.loads(f.read())
    f.close()
    for item in user_config: config[item] = user_config[item]
    os.remove("update.txt")
    save_config()
except:
    pass

session = {}

def submission_date(ID):
    '''This function returns the date the Qualtrics survey was submitted (the EndDate field).'''
    if ID not in qualtrics: return(None)
    if "EndDate" in qualtrics[ID]:
        try:
            return(datetime.strptime(qualtrics[ID]["EndDate"], '%Y-%m-%d %H:%M:%S'))
        except:
            print("Warning: Cannot parse the 'EndDate' field. This error sometimes occurs when the Qualtrics file has been modified in Excel.")
            return(None)
    else:
        return(None)

def days_late(ID, weekdays_only = True):
    '''This function returns the number of days late an assignment was (or 0 if it was submitted before time). By default, it only counts weekdays.'''
    sub_date = submission_date(ID)
    if "DueDate" in session and type(sub_date) is datetime:
        due_date = session["DueDate"]
        sub_date = datetime(sub_date.year, sub_date.month, sub_date.day, 0,0,0)
        due_date = datetime(due_date.year, due_date.month, due_date.day, 0,0,0)
        d_late = max(0, (sub_date - due_date).days)
        for a in range(0,d_late*weekdays_only):
            d_late -= ((sub_date - timedelta(days=a)).weekday() >= 5)
        return(d_late)
    else:
        return(None)

def ratings_given(ID):
    '''This function returns all of the ratings given by a rater (ID) in a list, including self-ratings'''
    if ID not in qualtrics: return([None])
    return(qualtrics[ID]["TeamWork"])

def rater_total(ID):
    '''A function that calculates the total number of points allocated by a rater to each peer'''
    if ID not in qualtrics: return([None])
    tot = [0] * (len(qualtrics[ID]["Peers"]) + 1) # Including self-rating
    for item in ratings_given(ID):
        for x, rating in enumerate(item):
            tot[x] = tot[x] + rating
    return(tot)

def rater_peerID(ID):
    ''' Returns the IDs of the peer raters of a person in a list using 2 methods'''
    tot = []
    if ID in qualtrics: # Method 1 - look up the person's record
        for peer in qualtrics[ID]["Peers"]: tot = tot + [int(peer)]
    else: # Method 2 - Look through all of their peers and grab IDs  
        for user in qualtrics:
            if "Peers" in qualtrics[user]:
                if ID in qualtrics[user]["Peers"]:
                    tot.append(int(user))
    tot.insert(0, int(ID))
    return(tot)

def max_peers():
    '''A function that shows the largest number of peers'''
    maxp = 0
    for ID in orig_data:
        len_p = len(rater_peerID(int(ID)))
        if len_p > maxp: maxp = len_p
    return(maxp)

def NonComplete(ID):
    '''A function to detect whether a person has not completed the peer survey'''
    # Remember, we remove all students who have not filled out any of the TeamWork items
    return(not(ID in qualtrics))

def PartialComplete(ID):
    '''A function to detect whether a person has only partially completed the peer survey'''
    n_peers = len(rater_peerID(ID))
    for item in qualtrics[ID]["TeamWork"]:
        if len(item) < n_peers: return(True)
    return(False)

def SelfPerfect(ID):
    '''A function to detect whether a person has given themselves all top ratings'''
    if NonComplete(ID) == True: return(False)
    for item in ratings_given(ID):
        if len(item) > 0:
            if item[0] < config["PointsPossible"] - 1: return(False)    
    return(True)

def PeersAllZero(ID):
    '''A function to detect whether a person has given their peers all zeroes. It assumes that the minimum number of peers has been met.'''
    if len(qualtrics[ID]["Peers"]) < config["MinimumPeers"]: return(False)
    if NonComplete(ID) == True: return(False)
    for item in ratings_given(ID):
        for rating in item[1:]:
            if rating > 0: return(False)
    return(True)


def ratings_received(ID):
    '''A function that returns all of the ratings received by a person (ratings nested within items)'''
    rr = []
    for y in tw_list: rr.append([])
    for x, rater_ID in enumerate(rater_peerID(ID)):
        # Only include the ratings if the person: (1) is in the Qualtrics file, (2) is not excluded because of partial completes, perfect self scores, or rating peers as all zeroes
        peer_ratings = ratings_given(rater_ID)
        for y in range(len(rr)):
            if rater_ID in qualtrics and not(
                    (config["Exclude_PartialComplete"] == 1 and PartialComplete(rater_ID) == True) or
                    (config["Exclude_SelfPerfect"] == 1 and SelfPerfect(rater_ID) == True) or
                    (config["Exclude_PeersAllZero"] == 1 and PeersAllZero(rater_ID) == True) or
                     len(peer_ratings[y])==0):
                rr[y].append(peer_ratings[y][rater_peerID(rater_ID).index(ID)])
            else:
                rr[y].append(None)
    return(rr)
    
def ratee_total(ID):
    '''A function that calculates the total number of points received from each peer, including oneself.'''
    tot = []
    rr = ratings_received(ID)
    for x, rater in enumerate(rater_peerID(ID)):
        tot.append(None)
        for y, item in enumerate(rr):
            if item[x] is not None:
                if tot[x] is None: tot[x] = 0
                tot[x] += item[x]
    return(tot)

def n_ratings(ID):
    '''A function that returns the number of valid ratings received from peers. It does not count the person's self-rating'''
    if len(rater_peerID(ID)) > 1:
        return(nansum([True if x is not None else False for x in ratee_total(ID)[1:]]))
    else:
        return(0)

def ratee_ratio(ID):
    '''A function that calculates the ratio of points received from a peer'''
    ratio = []
    for x, rater in enumerate(rater_peerID(ID)):
        if rater in qualtrics:
            ratings = rater_total(rater)
            ID_index = rater_peerID(rater).index(ID)
            if config["SelfVote"] not in [1,3]: ratings[0] = None # Only count the self-vote towards the ratio if self-voting allowed
            nratings = nansum([True if x is not None else False for x in ratings])
            if nansum(ratings) == 0: # Return NAN if the rater has given everyone a zero
                ratio.append(None)
            elif ID == rater: # Only calculate the ratio for one's own score if self-scoring allowed
                if config["SelfVote"] in [1,3]:
                    ratio.append(None)
                else:
                    if ratings[ID_index] is None:
                        ratio.append(None)
                    else:
                        ratio.append(nratings * ratings[ID_index] / nansum(ratings))
            else: # Otherwise, return the ratio
                ratio.append(nratings * ratings[ID_index] / nansum(ratings))
        else: # Rater not found
            ratio.append(None)
    return(ratio)


def feedback(ID):
    '''A function to locate the feedback provided to each person'''
    fb = [[], []]
    for rater_ID in rater_peerID(ID)[1:]:
        if rater_ID in qualtrics:
            id_index = qualtrics[rater_ID]["Peers"].index(ID)
            fb[0].append(qualtrics[rater_ID]["Feedback1"][id_index])
            fb[1].append(qualtrics[rater_ID]["Feedback2"][id_index])
        else:
            fb[0].append(None)
            fb[1].append(None)
    return(fb)

def general_feedback(ID):
    '''A function that returns general feedback provided to the team'''
    gen_feed = [[], []]
    for rater_ID in rater_peerID(ID):
        if rater_ID in qualtrics:
            feed1 = qualtrics[rater_ID]["Feedback1"][-1]
            feed2 = qualtrics[rater_ID]["Feedback2"][-1]
            if len(feed1) > 0: gen_feed[0].append(feed1)
            if len(feed2) > 0: gen_feed[1].append(feed2)
    return(gen_feed)

def penalty(ID):
    '''A function to calculate the % deducted as a penalty'''
    pen = 1
    if NonComplete(ID): # Apply penalty if they haven't filled out the survey
        pen = pen - (config["Penalty_NonComplete"]/100)
    else:
        if PartialComplete(ID): pen = pen - (config["Penalty_PartialComplete"]/100)
        if SelfPerfect(ID): pen = pen - (config["Penalty_SelfPerfect"]/100)
        if PeersAllZero(ID): pen = pen - (config["Penalty_PeersAllZero"]/100)
        if days_late(ID, config["weekdays_only"]) > 0: pen = pen - (config["Penalty_PerDayLate"]/100)*days_late(ID, config["weekdays_only"])
    return(max(pen,0))

def adj_score(ID, apply_penalty = True):
    '''A function to calculate the adjusted score'''
    if ID in adj_dict: 
        orig_score = adj_dict[ID]
    else:
        if (config["ScoreType"] == 1): # Simply return the mean score the ratee received
            if n_ratings(ID) >= config["MinimumPeers"]:
                rat_total = ratee_total(ID)
                if config["SelfVote"]==0: rat_total[0] = None # Get rid of the self-rating
                if len(rater_total(ID)) > 1:
                    if config["SelfVote"]==2: rat_total[0] = nanmean(rater_total(ID)[1:]) # Automatic mean vote
                    if config["SelfVote"]==3: rat_total[0] = npmin([rat_total[0], nanmean(rater_total(ID)[1:])]) # Person cannot give themselves more than the average mark they have assigned others
                else:
                    rat_total[0] = None
                rat_mean = nanmean(rat_total) # Calculate the mean total score
                orig_score = rat_mean / len(tw_list) # Divide by the number of items
                if config["RescaleTo"] is not None and config["RescaleTo"] != 0 and type(orig_score) is not str: orig_score = orig_score*config["RescaleTo"]/(session["points_possible"]-1) # Rescale the score to the desired range
            else:
                orig_score = "Insufficient ratings received (" + str(n_ratings(ID)) + ")"
        elif (config["ScoreType"] == 2): # Return an adjusted scorem based on group mark
            try:
                orig_score = float(orig_data[str(ID)]["orig_grade"])
            except:
                return("The user's original score cannot be found.")
            ratio = ratee_ratio(ID)
            if config["SelfVote"] == 0: ratio[0] = None
            if config["SelfVote"] == 2: ratio[0] = 1 # For Option 2, automatically assign a vote of 1
            if config["SelfVote"] == 3: ratio[0] = min(1, ratio[0]) # For Option #3, do not allow self-votes to raise their score
            if n_ratings(ID) >= config["MinimumPeers"]: # Number of legit scores must exceed minimum peers
                orig_score = orig_score*nanmean(ratio)
            orig_score = min(orig_score, session["points_possible"]) # Score is not allowed to exceed number of points possible.
            if config["RescaleTo"] is not None and config["RescaleTo"] != 0 and type(orig_score) is not str: orig_score = orig_score*(config["RescaleTo"]/session["points_possible"]) # Rescale the score to the desired range
        adj_dict[ID] = orig_score
    if apply_penalty == True: 
        try:
            orig_score = orig_score*penalty(ID) # Apply the penalty if the person does not complete the peer evaluation
        except:
            pass
    return(orig_score)

# Comments - a function to provide the students with their scores on their peer ratings, as well as peer feedback
def comments(ID):
    comment = "Feedback is unavailable because not enough of your team members completed the peer evaluation.\n\n"
    penalty_text = ""
    if NonComplete(ID)==True:
        if config["Penalty_NonComplete"] > 0: penalty_text = penalty_text + "-" + str(config["Penalty_NonComplete"]) + "% because you did not complete the Peer Evaluation survey."
    else:
        if PartialComplete(ID) == True and config["Penalty_PartialComplete"] > 0: penalty_text = penalty_text + "-" + str(config["Penalty_PartialComplete"]) + "% because you only partially completed the Peer Evaluation survey.\n"
        if SelfPerfect(ID) == True and config["Penalty_SelfPerfect"] > 0: penalty_text = penalty_text + "-" + str(config["Penalty_SelfPerfect"]) + "% because you assigned yourself a perfect score on the Peer Evaluation survey across all questions.\n"
        if PeersAllZero(ID) == True and config["Penalty_PeersAllZero"] > 0: penalty_text = penalty_text + "-" + str(config["Penalty_PeersAllZero"]) + "% because you gave all of your peers the bottom score on the Peer Evaluation survey across all questions.\n"
        if days_late(ID, config["weekdays_only"]) is not None:
            if days_late(ID, config["weekdays_only"]) > 0 and config["Penalty_PerDayLate"] > 0: penalty_text = penalty_text + "-" + str(config["Penalty_PerDayLate"] * days_late(ID, config["weekdays_only"])) + "% because you submitted your evaluation " + str(days_late(ID, config["weekdays_only"])) + " day(s) late.\n"
    if len(penalty_text) > 0: comment = comment + "The following penalties have been subtracted from your Peer Mark:\n" + penalty_text
    if n_ratings(ID) < config["MinimumPeers"]: return(comment)
    comment = ""
    general_comments = general_feedback(ID)[0]

    if len(general_comments) > 0:
        comment = comment + "The following general comments about your team (not about yourself personally) were provided by your peers:\n\n"
        shuffle(general_comments) # Shuffle the general comments so the rater is not identifiable
        for temp_comment in general_comments:
            comment = comment + '"' + temp_comment + '"\n\n'
    peer_feedback = remove_blanks(feedback(ID)[0])
    if len(peer_feedback) > 0:
        comment = comment + "You have received the following specific feedback from your peers:\n\n"
        shuffle(peer_feedback) # Shuffle the order of the feedback so the rater is not identifiable
        for feed in peer_feedback: comment = comment + '"' + str(feed) + '"\n\n'
    comment = comment + "You have received the following ratings from your peers:\n\n"
    rr = ratings_received(ID)
    for y, TeamWork in enumerate(tw_names):
        comment = comment + TeamWork + "\n"
        self_rating = rr[y][0]
        peer_rating = nanmean(rr[y][1:])
        if self_rating is not None: comment = comment + "Your Self-Rating: {}/5\n".format(self_rating+1)
        if peer_rating is not None: comment = comment + "Average Rating Received from Peers: {}/5\n\n".format(round(peer_rating,2)+1)
    if config["ScoreType"] == 2:
        orig_score = float(orig_data[str(ID)]["orig_grade"])
        comment = comment + "Your mark was calculated using the following formula. "
        comment = comment + "You received " + str(round(orig_score,2)) + " out of " + str(session["points_possible"]) + " (" + str(round(100 * orig_score / session["points_possible"], 1)) + "%) for your team assignment. "
        comment = comment + "Based on this, you were assigned an initial mark of " + str(round(config["RescaleTo"] * orig_score / session["points_possible"],1)) + " out of " + str(config["RescaleTo"]) + " ("+ str(round(100 * orig_score / session["points_possible"], 1)) + "%). "
        comment = comment + "You received peer ratings that were "
        ratio = ratee_ratio(ID)
        if config["SelfVote"] == 0: ratio[0] = None
        if config["SelfVote"] == 2: ratio[0] = 1 # For Option 2, automatically assign a vote of 1
        if config["SelfVote"] == 3: ratio[0] = min(1, ratio[0]) # For Option #3, do not allow self-votes to raise their score
        ratio = nanmean(ratio)
        a_score = ratio * config["RescaleTo"] * orig_score / session["points_possible"]
        ratio = ratio - 1
        ratio = round(ratio * 100,1)
        if ratio < 0: comment = comment + str(abs(ratio)) + "% lower than"
        if ratio == 0: comment = comment + "equal to"
        if ratio > 0: comment = comment + str(ratio) + "% higher than"
        comment = comment + " the average rating received by your other team members. "
        if ratio != 0:
            comment = comment + "Your initial mark was therefore adjusted from " + str(round(config["RescaleTo"] * orig_score / session["points_possible"],2)) + " to your final Peer Mark of " + str(round(a_score, 2)) + ".\n\n"
        else: 
            comment = comment + "Your initial mark was therefore not adjusted.\n\n"
    if len(penalty_text) > 0: comment = comment + "Your Peer Mark has received the following penalties:\n" + penalty_text
    return(comment)

def cleanupURL(URL):
    URL = URL.strip().lower()
    if URL[-1] != "/": URL += "/"
    URL = URL.replace('http:', 'https:')
    return(URL)

def writeCSV(data, filename):
    txt = []
    for x, field in enumerate(data.keys()):
        if x == 0:
            n_cases = len(data[field])
        else:
            txt.append(",")
            if len(data[field]) != n_cases:
                messagebox.showinfo("Error", "Cannot write CSV file. The number of rows appears to be different.")
                sys.exit()
        txt.append("\"{}\"".format(field))
    txt.append("\n")
    for y in range(n_cases):
        for x, field in enumerate(data.keys()):
            if x > 0: txt.append(",")
            if "{}".format(data[field][y]) == "nan":
                txt.append("")
            elif type(data[field][y]) in [int, float]:
                txt.append("{}".format(data[field][y]))
            elif type(data[field][y]) == str:
                txt.append("\"{}\"".format(data[field][y]))
            else:
                txt.append("")
        txt.append("\n")
    f = open(filename, "a")
    f.write("".join(txt[0:-1]))
    f.close()

def writeXL(filename, datasets = [], sheetnames = []):
    wb = Workbook()
    for z in range(len(datasets)):
        if z==0:
            ws = wb.active
            ws.title = sheetnames[0]
        else:
            ws = wb.create_sheet(sheetnames[z])
        for x, field in enumerate(datasets[z].keys(),1):
            ws.cell(row = 1, column = x).value = clean_text(field)
            for y, temp_val in enumerate(datasets[z][field],2):
                ws.cell(row = y, column = x).value = clean_text(temp_val)
        ws.auto_filter.ref = 'A1:' + get_column_letter(ws.max_column) + str(ws.max_row)
        ws.freeze_panes = ws["A2"]
    wb.save(filename)

def cleanupNumber(txt):
    if type(txt) == str:
        txt = txt.strip()
        try:
            return int(txt)
        except ValueError:
            try:
                if int(float(txt)) == float(txt): return(int(float(txt)))
                return(float(txt))
            except:
                return(txt)
    else:
        return(txt)

def readTable(filename, delimiter = ","):
    f = open(filename)
    readFile = csv.reader((x.replace('\0','').replace('\r',"") for x in f), delimiter = delimiter)
    df = {}
    for y, row in enumerate(readFile,1):
        if len(row) > 0:
            if y == 1:
                for field in row:
                    df[field] = []
            else:
                for x, field in enumerate(df.keys()):
                    df[field].append(cleanupNumber(row[x]))
    f.close()
    return(df)

def readXL(filename):
    df = {}
    wb = load_workbook(filename)
    sheet_obj = wb.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    for x in range(1, max_col+1):
        col_name = sheet_obj.cell(row = 1, column = x).value
        df[col_name] = []
        for y in range(2, max_row+1):
            df[col_name].append(sheet_obj.cell(row = y, column = x).value)
    return(df)

def df_byRow(df, id_col):
    '''This function converts the data frame into a records format, such that it is organised via an id number'''
    df2 = {}
    for key in df[id_col]:
        df2[cleanupNumber(key)] = {} # Converts to a number if possible
    for x, col in enumerate(df):
        if col != id_col:
            for y, row in enumerate(df[col]):
                df2[df[id_col][y]][col] = cleanupNumber(df[col][y])
    return(df2)

def collapseList(df, list_var, vars, begin = 1, delete_blank = True):
    '''This function collapses a set of dictionary elements into a list'''
    if type(vars) == list:
        for case in df:
            df[case][list_var] = []
            for var in vars:
                if var in df[case]:
                    if df[case][var] != "" and type(df[case][var]) != type(None):
                        df[case][list_var].append(cleanupNumber(df[case].pop(var))) # Convert to a number if possible; otherwise, use string.
                    else:
                        if delete_blank == False: df[case][list_var].append("")
                        del df[case][var]
    elif type(vars) == str:
        if "*" in vars:
            for case in df:
                df[case][list_var] = []
                a = begin
                var = vars.replace("*", str(a))
                while var in df[case]:
                    if df[case][var] != "" and type(df[case][var]) != type(None):
                        df[case][list_var].append(cleanupNumber(df[case].pop(var)))
                    else:
                        if delete_blank == False: df[case][list_var].append("")
                        del df[case][var]
                    a = a + 1
                    var = vars.replace("*", str(a))
    return(df)

def df_byCol(df, id_col = "ExternalDataReference"):
    df2 = {id_col: []}
    for case in df:
        for col_name in df[case].keys():
            if type(df[case][col_name]) is list: # Parse a list as different columns
                for x in range(len(df[case][col_name])):
                    col_name2 = "{}_{}".format(col_name, x+1)
                    if col_name2 not in df2: df2[col_name2] = []
            else:
                if col_name not in df2: df2[col_name] = []
        
    for case in df.keys():
        for col_name in df2.keys(): df2[col_name].append(None) # Add blank values by default
        df2[id_col].append(case) # Add case number
        for col_name in df[case].keys():
            if type(df[case][col_name]) is list:
                for x in range(len(df[case][col_name])):
                    df2["{}_{}".format(col_name,x+1)][-1] = df[case][col_name][x]
            else:
                df2[col_name][-1] = df[case][col_name]
    return(df2)

def flatten_list(lst):
    lst2 = []
    for item in lst:
        if type(item) is list:
            lst2 = lst2 + flatten_list(item)
        elif item is None or item is str:
            pass
        else:
            lst2 = lst2 + [item]
    return(lst2)

def nansum(lst):
    return(sum(flatten_list(lst)))
    
def nanmean(lst):
    lst2 = flatten_list(lst)
    return(sum(lst2)/len(lst2))

def npmin(lst):
    lst2 = flatten_list(lst)
    return(min(lst2))

#%% GraphQL
class GraphQL():
    def __init__(self, course_id = 0, groupset_id = 0):
        self.course_id = course_id
        self.groupset_id = groupset_id
        self.cache = {}
    
    def query(self, query_text):
        global mm
        query = {'access_token': config["API_TOKEN"], 'query': query_text}
        for attempt in range(5):
            try:
                result = requests.post(url = "{}api/graphql".format(config["API_URL"]), data = query).json()
            except:
                print("Trying to connect to Canvas via GraphQL. Attempt {} of 5.".format(attempt+1))
            else:
                return(result)
        messagebox.showinfo("Error", "Cannot connect to Canvas via GraphQL. There may be an Internet connectivity problem or the Canvas base URL may be incorrect.")
        Settings_call()
    
    def courses(self):
        if "courses" in self.cache: return(self.cache["courses"])
        temp = self.query('query MyQuery {allCourses {_id\nname}}')["data"]["allCourses"]
        temp2 = {}
        for course in temp:
            temp2[int(course["_id"])] = {}
            temp2[int(course["_id"])]["name"] = course["name"]
        df = {}
        for a in sorted(temp2.keys(), reverse = True):
            df[str(a)] = temp2[a]
        self.cache["courses"] = df
        return(df)

    def group_sets(self):
        if "groupset_" + str(self.course_id) not in self.cache:
            self.cache["groupset_" + str(self.course_id)] = self.query('query MyQuery {course(id:"' + str(self.course_id) + '") {groupSetsConnection {nodes {_id\nname}}}}')["data"]["course"]["groupSetsConnection"]["nodes"]
        return(self.cache["groupset_" + str(self.course_id)])

    def sections(self):
        if "sections_" + str(self.course_id) not in self.cache:
            query_result = self.query('query MyQuery {course(id: "' + str(self.course_id) + '") {sectionsConnection {nodes {_id\nname}}}}')["data"]["course"]["sectionsConnection"]["nodes"]
            self.cache["sections_" + str(self.course_id)] = {}
            for section in query_result:
                self.cache["sections_" + str(self.course_id)][section["_id"]] = {}
                self.cache["sections_" + str(self.course_id)][section["_id"]]["name"] = section["name"]
        return(self.cache["sections_" + str(self.course_id)])
    
    def students(self):
        if "students_" + str(self.course_id) not in self.cache:
            query_result = self.query('query MyQuery {course(id: "' + str(self.course_id) + '") {enrollmentsConnection {nodes {type\nuser {_id\nname\nemail}}}}}')
            self.cache["students_" + str(self.course_id)] = {"id": [], "name": [], "email": []}
            for student in query_result["data"]["course"]["enrollmentsConnection"]["nodes"]:
                if (student["type"] == "StudentEnrollment"):
                    if (student["user"]["_id"] not in self.cache["students_" + str(self.course_id)]["id"]):
                        self.cache["students_" + str(self.course_id)]["id"].append(student["user"]["_id"])
                        self.cache["students_" + str(self.course_id)]["name"].append(student["user"]["name"])
                        self.cache["students_" + str(self.course_id)]["email"].append(student["user"]["email"])
        return(self.cache["students_" + str(self.course_id)])
    
    def groups(self):
        if "groups_" + str(self.course_id) in self.cache: return(self.cache["groups_" + str(self.course_id)])
        query_result = self.query('query MyQuery {course(id: "' + str(self.course_id) + '") {groupSetsConnection {nodes {_id\ngroupsConnection {nodes {_id\nname\nmembersConnection {nodes {user {_id\nname}}}}}}}}}')
        try:
            for groupset in query_result["data"]["course"]["groupSetsConnection"]["nodes"]:
                if (groupset["_id"]==str(self.groupset_id)):
                    temp = groupset["groupsConnection"]["nodes"]
                    df = {}
                    for group in temp:
                        df[group["_id"]] = {"name": group["name"], "users": []}
                        for user in group["membersConnection"]["nodes"]:
                            df[group["_id"]]["users"].append(user["user"]["_id"])
                    self.cache["groups_" + str(self.course_id)] = df
                    return(df)
                    break
        except:
            return({})
        return({})
    
    def students_comprehensive(self):
        print("Course ID is: " + str(self.course_id))
        if "students_" + str(self.course_id) + "_" + str(self.groupset_id) in self.cache:
            return(self.cache["students_" + str(self.course_id) + "_" + str(self.groupset_id)])
        query_result = self.query('query MyQuery {course(id: "' + str(self.course_id) + '") {enrollmentsConnection {nodes {type\nuser {_id\nname\nemail}\nsection {_id}}}}}')["data"]["course"]["enrollmentsConnection"]["nodes"]
        student_series = {}
        # Set up dictionary for students
        for student in query_result:
            if (student["type"] == "StudentEnrollment"):
                if (student["user"]["_id"] not in student_series):
                    student_series[student["user"]["_id"]] = {"name": student["user"]["name"], "email": student["user"]["email"]}
                    student_series[student["user"]["_id"]]["first_name"] = student["user"]["name"][0:student["user"]["name"].find(" ")]
                    student_series[student["user"]["_id"]]["last_name"] = student["user"]["name"][student["user"]["name"].rfind(" ")+1:]
                    student_series[student["user"]["_id"]]["sections"] = []
                student_series[student["user"]["_id"]]["sections"].append(student["section"]["_id"])
        
        # Identify which group the student is in
        if self.groupset_id != 0: #Only look for groups if a group set has been specified
            groups = self.groups()
            def find_group(student, groups):
                for group in groups:
                    for member in groups[group]["users"]:
                        if (student == member):
                            return(group)
                return("0")
            for student in student_series:
                group = find_group(student, groups)
                if (group != "0"):
                    student_series[student]["group_id"] = group
                    student_series[student]["group_name"] = groups[group]["name"]
                            
        # Find the students' peers (if they are in a group)
        for student in student_series:
            if ("group_id" in student_series[student]):
                student_series[student]["peers"] = []
                for member in groups[student_series[student]["group_id"]]["users"]:
                    if (member != student):
                        student_series[student]["peers"].append(member)
                
        self.cache["students_" + str(self.course_id) + "_" + str(self.groupset_id)] = student_series                
        return(student_series)

    def teachers_comprehensive(self):
        if "teachers_" + str(self.course_id) in self.cache:
            return(self.cache["teachers_" + str(self.course_id)])
        query_result = self.query('query MyQuery {course(id: "' + str(self.course_id) + '") {enrollmentsConnection {nodes {type\nuser {_id\nname\nemail}\nsection {_id}}}}}')["data"]["course"]["enrollmentsConnection"]["nodes"]
        teacher_series = {}
        # Set up dictionary for teachers
        for teacher in query_result:
            if (teacher["type"] == "TeacherEnrollment"):
                if (teacher["user"]["_id"] not in teacher_series):
                    teacher_series[teacher["user"]["_id"]] = {"name": teacher["user"]["name"], "email": teacher["user"]["email"]}
                    teacher_series[teacher["user"]["_id"]]["first_name"] = teacher["user"]["name"][0:teacher["user"]["name"].find(" ")]
                    teacher_series[teacher["user"]["_id"]]["last_name"] = teacher["user"]["name"][teacher["user"]["name"].rfind(" ")+1:]
                    teacher_series[teacher["user"]["_id"]]["sections"] = []
                teacher_series[teacher["user"]["_id"]]["sections"].append(teacher["section"]["_id"])
                
        self.cache["teachers_" + str(self.course_id) + "_" + str(self.groupset_id)] = teacher_series                
        return(teacher_series)
        
        
#%% Tool Tips
class ToolTip(object):
# From: https://stackoverflow.com/questions/20399243/display-message-when-hovering-over-something-with-mouse-cursor-in-python
    
    def __init__(self, widget):
        self.widget = widget
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0

    def showtip(self, text):
        "Display text in tooltip window"
        self.text = text
        if self.tipwindow or not self.text:
            return
        x, y, cx, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 57
        y = y + cy + self.widget.winfo_rooty() +27
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(1)
        tw.wm_geometry("+%d+%d" % (x, y))
        label = tk.Label(tw, text=self.text, justify=tk.LEFT,
                      background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                      font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

def CreateToolTip(widget, text):
    toolTip = ToolTip(widget)
    def enter(event):
        toolTip.showtip(text)
    def leave(event):
        toolTip.hidetip()
    widget.bind('<Enter>', enter)
    widget.bind('<Leave>', leave)

#%% Settings
    
class Settings:
    def __init__(self, window):
        # Again, try to load the configuration file and overwrite defaults
        try:
            f = open("config.txt", "r")
            user_config = json.loads(f.read())
            f.close()
            for item in user_config: config[item] = user_config[item]
        except:
            pass
        
        self.window = window
        window.title("Settings")
        window.resizable(0, 0)
        window.grab_set()
        
        tk.Label(window, text = "\nPlease enter your API Key from Canvas:", fg = "black").grid(row = 2, column = 1, sticky = "W")

        tk.Label(window, text = "Get Token:", fg = "black").grid(row = 3, column = 0, sticky = "E")
        self.access_canvas = tk.Button(window, text = "Get Token from Canvas", fg = "black", command = self.open_canvas, state = "disabled")
        self.access_canvas.grid(row = 3, column = 1, sticky="W", padx = 5, pady = 5)
        CreateToolTip(self.access_canvas, text = 
              'Click on this button to access Canvas in your web browser.\n'+
              'From your profile page, click on the "+New Access Token" button.\n'+
              'Generate the token and copy/paste it into the box below.')

        tk.Label(window, text = "\nPlease enter the URL for the Canvas API: (e.g., https://swinburne.instructure.com/)", fg = "black").grid(row = 0, column = 1, sticky = "W")
        tk.Label(window, text = "API URL:", fg = "black").grid(row = 1, column = 0, sticky = "E")
        self.settings_URL = tk.Entry(window, width = 50)
        self.settings_URL.bind("<Key>", self.check_status)
        self.settings_URL.grid(row = 1, column = 1, sticky="W", padx = 5)
        CreateToolTip(self.settings_URL, text = 
              'The base URL used to log into Canvas.')

        tk.Label(window, text = "Enter Token:", fg = "black").grid(row = 4, column = 0, sticky = "E")
        self.settings_Token = tk.Entry(window, width = 80)
        self.settings_Token.grid(row = 4, column = 1, sticky="W", padx = 5)
        CreateToolTip(self.settings_Token, text = 
              'The Token allows PEER to log into Canvas on your behalf\n' +
              'and access information about students, groups and courses.')
        
        tk.Label(window, text = "You will only need to enter these settings once.", fg = "black").grid(row = 6, columnspan = 2)
        self.settings_openfolder = tk.Button(window, text = "Open App Folder", fg = "black", command = self.open_folder).grid(row = 7, columnspan = 2, pady = 5)
        self.settings_default = tk.Button(window, text = "Restore Default Settings", fg = "black", command = self.restore_defaults).grid(row = 8, columnspan = 2, pady = 5)
        self.settings_save = tk.Button(window, text = "Save Settings", fg = "black", command = self.save_settings).grid(row = 9, columnspan = 2, pady = 5)
        self.update_fields()
        
        self.check_status()

    def open_canvas(self):
        open_url(cleanupURL(self.settings_URL.get()) + "profile/settings")

    def open_folder(self):
        subprocess.Popen(['explorer', os.path.abspath(os.path.dirname(sys.argv[0]))])
        self.window.destroy()

    def update_fields(self):
        self.settings_URL.delete(0, tk.END)
        self.settings_URL.insert(0, config["API_URL"])
        self.settings_Token.delete(0, tk.END)
        self.settings_Token.insert(0, config["API_TOKEN"])
        
    def restore_defaults(self):
        global config
        config = {}
        for item in defaults: config[item] = defaults[item]
        self.update_fields()
        
    def save_settings(self):
        config["API_URL"] = cleanupURL(self.settings_URL.get())
        config["API_TOKEN"] = self.settings_Token.get().strip()
        save_config()
        self.window.destroy()
        
    def check_status(self, key_press = ""):
        txt = self.settings_URL.get().lower()
        if type(key_press) is not str: txt = txt + key_press.char.lower()
        if len(txt) >= 17 and txt[-17:] == ".instructure.com/" and (txt[0:7]=="http://" or txt[0:8]=="https://"):
            self.access_canvas["state"] = "normal"
        else:
            self.access_canvas["state"] = "disabled"

def Settings_call():
    settings_win = tk.Toplevel()
    set_win = Settings(settings_win)

#%% Export Groups
class ExportGroups():
    def __init__(self, master, preset = ""):
        self.preset = preset
        self.master = master
        self.exportgroups = tk.Frame(self.master)
        self.exportgroups.pack()
        
        if preset=="classlist":
            self.master.title("Download Class List")
            tk.Label(self.exportgroups, text = "\nDownload list of students and teachers", fg = "black").grid(row = 0, column = 0, columnspan = 2)            
        else:
            self.master.title("Create Contact List")
            tk.Label(self.exportgroups, text = "\nDownload student groups and group membership", fg = "black").grid(row = 0, column = 0, columnspan = 2)

        self.master.resizable(0, 0)
        self.master.grab_set()

        tk.Label(self.exportgroups, text = "Unit:", fg = "black").grid(row = 1, column = 0, sticky = "E", pady = 5)           
        self.exportgroups_unit = ttk.Combobox(self.exportgroups, width = 60, values = [], state="readonly")
        self.exportgroups_unit.bind("<<ComboboxSelected>>", self.group_sets)
        self.exportgroups_unit.grid(row = 1, column = 1, sticky = "W", padx = 5)

        tk.Label(self.exportgroups, text = "Group Set:", fg = "black").grid(row = 2, column = 0, sticky = "E", pady = 5)
        self.exportgroups_groupset = ttk.Combobox(self.exportgroups, width = 60, state="readonly")
        self.exportgroups_groupset.grid(row = 2, column = 1, sticky = "W", padx = 5)
        
        self.exportgroups_export = tk.Button(self.exportgroups, text = "Start Download", fg = "black", command = self.begin_export)
        self.exportgroups_export.grid(row = 4, columnspan = 2, pady = 5)

        self.group_sets_id = []
        self.group_sets_name = []          
        
        self.exportgroups_unit["values"] = tuple(session["course.names"])
        self.exportgroups_unit.current(0)
        self.group_sets()
        
        self.statusbar = tk.Label(self.exportgroups, text="", bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.statusbar.grid(row = 5, columnspan = 2, sticky = "ew")
        
        self.status("")
        
        if preset!="classlist" and config["firsttime_export"] == True:
            messagebox.showinfo("Reminder", "Before creating a Contacts list for Qualtrics, it is recommended you check the accuracy of the group membership data in Canvas. Use the \"Confirm group membership\" function in the main menu to contact students to confirm their membership in each team.")
            config["firsttime_export"] = False
            save_config()
    
    def group_sets(self, event_type = ""):
        self.group_sets_id = []
        self.group_sets_name = []
        GQL.course_id = session["course.ids"][self.exportgroups_unit.current()]
        for group_set in GQL.group_sets():
            self.group_sets_id.append(group_set["_id"])
            self.group_sets_name.append(group_set["name"])
        if len(self.group_sets_id) > 0:
            self.exportgroups_groupset["values"] = tuple(self.group_sets_name)
            self.exportgroups_groupset.current(0)
        else:
            self.exportgroups_groupset["values"] = ()
            self.exportgroups_groupset.set('')
    
    def status(self, text): 
        self.statusbar["text"] = text
        self.statusbar.update_idletasks()
    
    def begin_export(self, event_type = ""):
        if self.preset=="classlist":
            filename = tk.filedialog.asksaveasfilename(initialfile=session["course.names"][self.exportgroups_unit.current()], defaultextension=".xlsx", title = "Save Exported Data As...", filetypes = (("Excel files","*.xlsx"),("all files","*.*")))
        else:
            filename = tk.filedialog.asksaveasfilename(initialfile="Contacts List", defaultextension=".csv", title = "Save Contacts List As...", filetypes = (("Comma separated values files","*.csv"),("all files","*.*")))

        GQL.course_id = session["course.ids"][self.exportgroups_unit.current()]
        if self.exportgroups_groupset.current() >= 0:
            GQL.groupset_id = self.group_sets_id[self.exportgroups_groupset.current()]
            session["group_set"] = self.group_sets_id[self.exportgroups_groupset.current()]    
        session["course.id"] = session["course.ids"][self.exportgroups_unit.current()]
        session["course.name"] = session["course.names"][self.exportgroups_unit.current()]
        print("Accessing " + session["course.name"] + "...")        
        students = GQL.students_comprehensive()
        print("Found {} students...".format(len(students)))

        if self.preset=="classlist":
            sections = GQL.sections()
            teachers = GQL.teachers_comprehensive()
            print("Found {} teachers...".format(len(teachers)))
            
            student_list = {"id": [],
                         "name":[],
                         "first_name":[],
                         "last_name":[],
                         "email":[],
                         "group_name":[]}
            
            teacher_list = {"id": [],
                         "name":[],
                         "first_name":[],
                         "last_name":[],
                         "email":[]}
            
            for section in sections:
                student_list[sections[section]["name"]] = []
                teacher_list[sections[section]["name"]] = []
            for student in students:
                student_list["id"].append(student)
                student_list["name"].append(students[student]["name"])
                student_list["first_name"].append(students[student]["first_name"])
                student_list["last_name"].append(students[student]["last_name"])
                student_list["email"].append(students[student]["email"])
                if "group_name" in students[student]:
                    student_list["group_name"].append(students[student]["group_name"])
                else:
                    student_list["group_name"].append("")
                for section in sections:
                    if section in students[student]["sections"]:
                        student_list[sections[section]["name"]].append(True)
                    else:
                        student_list[sections[section]["name"]].append(False)
            
            for teacher in teachers:
                teacher_list["id"].append(teacher)
                teacher_list["name"].append(teachers[teacher]["name"])
                teacher_list["first_name"].append(teachers[teacher]["first_name"])
                teacher_list["last_name"].append(teachers[teacher]["last_name"])
                teacher_list["email"].append(teachers[teacher]["email"])                
                for section in sections:
                    if section in teachers[teacher]["sections"]:
                        teacher_list[sections[section]["name"]].append(True)
                    else:
                        teacher_list[sections[section]["name"]].append(False)
            
            # New code to write to XLSX
            writeXL(filename=filename, datasets=[student_list, teacher_list], sheetnames=["Student List", "Teacher List"])
            messagebox.showinfo("Download complete.", "Student list saved as {}.".format(filename))
            
        else:
            student_list = {"ExternalDataReference":[],
                            "Student_name":[],
                            "FirstName":[],
                            "LastName":[],
                            "Email":[],
                            "Group_id":[], "Group_name":[],
                            "Peer1_id":[], "Peer1_name":[],
                            "Peer2_id":[], "Peer2_name":[],
                            "Peer3_id":[], "Peer3_name":[],
                            "Peer4_id":[], "Peer4_name":[],
                            "Peer5_id":[], "Peer5_name":[],
                            "Peer6_id":[], "Peer6_name":[],
                            "Peer7_id":[], "Peer7_name":[],
                            "Peer8_id":[], "Peer8_name":[],
                            "Peer9_id":[], "Peer9_name":[]}
            
            print("Processing student details...")
            for student in students:
                student_list["ExternalDataReference"].append(student)
                student_list["Student_name"].append(students[student]["name"])
                student_list["FirstName"].append(students[student]["first_name"])
                student_list["LastName"].append(students[student]["last_name"])
                student_list["Email"].append(students[student]["email"])
                if ("group_id" in students[student]):
                    print("> {}: {}".format(students[student]["name"], students[student]["group_name"]))
                    student_list["Group_id"].append(students[student]["group_id"])            
                    student_list["Group_name"].append(students[student]["group_name"])
                    for j, peer in enumerate(students[student]["peers"],1):
                        student_list["Peer" + str(j) + "_id"].append(peer)
                        student_list["Peer" + str(j) + "_name"].append(students[peer]["name"])
                else:
                    print("> {}: No group found".format(students[student]["name"]))
                    student_list["Group_id"].append(None)            
                    student_list["Group_name"].append(None)
                    j = 0
                for k in range(j+1, 10):
                    student_list["Peer" + str(k) + "_id"].append(None)
                    student_list["Peer" + str(k) + "_name"].append(None)
                            
            # Write CSV of group members
            print("Creating Contacts file to be uploaded to Qualtrics")
            writeCSV(student_list, filename)
            messagebox.showinfo("Download complete.", "Qualtrics Contact file saved as \"{}\".".format(filename))
        
        self.master.destroy()

def ExportGroups_call():
    exportgroups = tk.Toplevel()
    exp_grp = ExportGroups(exportgroups)

def ExportGroups_classlist():
    exportgroups = tk.Toplevel()
    exp_grp = ExportGroups(exportgroups, preset="classlist")

#%% Peer Mark section
class PeerMark():
   def __init__(self, master):
       self.master = master
       self.master.title("Calculate Peer Marks")
       self.master.resizable(0, 0)
       
       self.peermark = tk.Frame(self.master)
       self.peermark.pack()
       
       #self.sentinel = None
    
       self.UploadPeerMark = tk.BooleanVar()
       self.UploadPeerMark.set(True)
    
       tk.Label(self.peermark, text = "Unit:", fg = "black").grid(row = 0, column = 0, sticky = "E", pady = 5)
       self.peermark_unit = ttk.Combobox(self.peermark, width = 60, state="readonly")
       self.peermark_unit.bind("<<ComboboxSelected>>", self.group_assns)
       self.peermark_unit.grid(row = 0, column = 1, sticky = "W", padx = 5)
       CreateToolTip(self.peermark_unit, text =
                      'Select the unit that you want to upload the peer marks to.')      
    
       tk.Label(self.peermark, text = "Team assessment task:", fg = "black").grid(row = 1, column = 0, sticky = "E", pady = 5)
       self.peermark_assn = ttk.Combobox(self.peermark, width = 60, state="readonly")
       self.peermark_assn.grid(row = 1, column = 1, sticky = "W", padx = 5)
       CreateToolTip(self.peermark_assn, text =
                      'If you are using the adjusted scoring method, select the team\n'+
                      'assignment used to calculate the student\'s baseline peer mark.')
    
       tk.Label(self.peermark, text = "Qualtrics file:", fg = "black").grid(row = 2, column = 0, sticky = "E", pady = 5)
       self.peermark_peerdata = tk.Button(self.peermark, text = "Select...", fg = "black", command = self.select_file)
       self.peermark_peerdata.grid(row = 2, column = 1, padx = 5, pady = 5, sticky = "W")
       self.datafile = tk.Label(self.peermark, text = "", fg = "black")
       self.datafile.grid(row = 3, column = 1, pady = 5, sticky = "W")
       CreateToolTip(self.peermark_peerdata, text =
                      'Select a Qualtrics file containing the survey response data.')
    
       tk.Label(self.peermark, text = "Peer evaluation due date:", fg = "black").grid(row = 4, column = 0, sticky = "E")
       self.peermark_duedate = DateEntry(self.peermark, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern = "yyyy-mm-dd")
       self.peermark_duedate.grid(row = 4, column = 1, padx = 5, sticky = "W")
       CreateToolTip(self.peermark_duedate, text =
                      'The due date for the Peer Evaluation survey. Late penalties\n' +
                      'can be applied to students submitting after this date.')

       tk.Label(self.peermark, text = "Scoring policies:", fg = "black").grid(row = 5, column = 0, sticky = "E")    
       self.peermark_customise = tk.Button(self.peermark, text = "Review/modify", fg = "black", command = Policies_call)
       self.peermark_customise.grid(row = 5, column = 1, padx = 5, pady = 5, sticky = "W")
       CreateToolTip(self.peermark_customise, text =
                      'Review/modify the policies for scoring the peer mark.')
    
       self.peermark_calculate = tk.Button(self.peermark, text = "Calculate marks", fg = "black", command = self.start_calculate, state = 'disabled')
       self.peermark_calculate.grid(row = 6, column = 0, padx = 5, columnspan = 2, pady = 5)
       CreateToolTip(self.peermark_calculate, text =
                      'Calculate the peer marks and upload them to Canvas, along\n' +
                      'with peer feedback.')
       
       self.peermark_unit["values"] = tuple(session["course.names"])
       self.peermark_unit.current(0)
       self.group_assns()
        
       self.status_text = tk.StringVar()
       self.statusbar = tk.Label(self.peermark, bd=1, relief=tk.SUNKEN, anchor=tk.W, textvariable = self.status_text)
       self.statusbar.grid(row = 7, columnspan = 2, sticky = "ew")       
       
       self.master.bind("<FocusIn>", self.check_status)
       
       if config["firsttime_upload"] == True:
           messagebox.showinfo("Reminder", "Before uploading the peer evaluation marks to Canvas, it is recommended that you first review the scoring policies. Click on the \"Review/modify\" button to review these settings.")
           config["firsttime_upload"] = False
           save_config()

   def select_file(self):
       session["DataFile"] = filedialog.askopenfilename(initialdir = ".",title = "Select file",filetypes = (("Compatible File Types",("*.csv","*.tsv","*.xlsx")),("All Files","*.*")))
       if len(session["DataFile"]) > 0:
           self.peermark_calculate["state"] = 'normal'
       else:
           self.peermark_calculate["state"] = 'disabled'
       self.datafile["text"] = session["DataFile"]

   def group_assns(self, event_type = ""):
        self.assn_id = []
        self.assn_name = []
        course = canvas.get_course(session["course.ids"][self.peermark_unit.current()])
        assns = course.get_assignments()
        for assn in assns:
            self.assn_id.append(assn.id)
            self.assn_name.append(assn.name)
        self.peermark_assn["values"] = tuple(self.assn_name)
        self.peermark_assn.current(0)

   def check_status(self, event_type):
       self.master.grab_set()    
       if config["ScoreType"] == 1:
           self.peermark_assn["state"] = "disabled"
       else:
           self.peermark_assn["state"] = "readonly"
    
   def status(self, stat_text):
       self.status_text.set(stat_text)
       
   def start_calculate(self):
       self.status("Loading, please be patient...")
       session["DueDate"] = self.peermark_duedate.get_date()
       session["GroupAssignment_ID"] = self.assn_id[self.peermark_assn.current()]
       GQL.course_id = session["course.ids"][self.peermark_unit.current()]
       GQL.groupset_id = 0
       session["course.name"] = session["course.names"][self.peermark_unit.current()]
       course = canvas.get_course(session["course.ids"][self.peermark_unit.current()])
       calculate_marks(course, self)

def PeerMark_call():
    global pm
    pm = tk.Toplevel()
    window = PeerMark(pm)

#%% Calculate marks   
def calculate_marks(course, pm): 
   if config["ExportData"]:
       filename = ""
       filename = tk.filedialog.asksaveasfilename(initialfile=session["course.name"] + " (Peer Data Export)", defaultextension=".xlsx", title = "Save Exported Data As...", filetypes = (("Excel files","*.xlsx"),("all files","*.*")))
    
   print("Calculating peer marks... please be patient.")
   global qualtrics, orig_data, tw_list, tw_names, other_list, ratio_dict, adj_dict
   
   # Set up data frame
   orig_data = GQL.students_comprehensive()
   other_list = [] # Flush the "other" list.
   
   # Load the Qualtrics file
   print("Loading the Qualtrics file...")
   try:
       if session["DataFile"][-3:].upper()=="CSV":
           qualtrics = readTable(session["DataFile"])
       elif session["DataFile"][-3:].upper() in ["TSV","TXT"]:
           qualtrics = readTable(session["DataFile"], delimiter="\t")
       elif session["DataFile"][-3:].upper() == "XLS" or session["DataFile"][-4:].upper() in ["XLSX", "XLSM"]:
           qualtrics = readXL(session["DataFile"]) #load_workbook(session["DataFile"])
       else:
           messagebox.showinfo("Error", "Unable to import data from this file type.")
           pm.master.destroy()
   except:
       messagebox.showinfo("Error", "Unable to parse data from this file.")
       pm.master.destroy()
         
   # Create new assignment
   if config["UploadData"]:
       print("Creating new assignment for Peer Mark")
       new_assignment = course.create_assignment({'name': config["PeerMark_Name"],
                                                  'notify_of_update': False,
                                                  'points_possible': config["RescaleTo"],
                                                  'description': 'This assignment is your Peer Mark for the team project.',
                                                  'published': True,
                                                  'due_at': datetime.strftime(session["DueDate"], '%Y-%m-%dT23:59:59')})
         
       print("New assignment created: " + new_assignment.name + " (" + str(new_assignment.id) + ")")
       
       # Set the assignment to manually reveal grades to students (i.e., hide the grades until released), using GraphQL
       query = {'access_token': config["API_TOKEN"],
                'query': 'mutation MyMutation {\nsetAssignmentPostPolicy(input: {assignmentId: ' + str(new_assignment.id) + ', postManually: true}) {\npostPolicy {\npostManually\n}\n}}'
                }
       url = "{}api/graphql".format(config["API_URL"])
       requests.post(url, data = query)

   # Get assignment details    
   if (config["ScoreType"]==1):
       # Use this method if we are only calculating totals. Download all users, but not submission score.
       print("Downloading list of students...")
       session["points_possible"] = config["PointsPossible"]
       for user in orig_data:
           orig_data[str(user)]["orig_grade"] = None
   elif (config["ScoreType"]==2):
       # Download student details and grades and submission score
       assignment = course.get_assignment(session["GroupAssignment_ID"])    
       print("Downloading group marks for " + assignment.name)
       session["points_possible"] = assignment.points_possible
       submissions = assignment.get_submissions()
       for submission in submissions:
           user = submission.user_id
           if str(user) in orig_data: # Leave out the Test Student
               print("{} ({}): {}".format(orig_data[str(user)]["name"], user, submission.score))
               orig_data[str(user)]["orig_grade"] = submission.score
   
   # Clean up Qualtrics
   print("Cleaning up Qualtrics data file...")
   rep_list = ["Please rate each team member in relation to: ", " - Myself"]
   for a in range(1,10): rep_list.append(" - [Field-Peer" + str(a) + "_name]")
   
   # Determine the teamwork items in the Qualtrics file
   for column in qualtrics:
       if (column[0:8].upper() == "TEAMWORK"):
           t_name = qualtrics[column][0]
           t_name.replace("&shy", "\u00AD")
           if (column.find("_") != -1):
               tw_list.append(column[0:column.rfind("_")])
               if t_name.count("\u00AD") == 2:
                   str_loc = find_all(t_name, "\u00AD")
                   tw_names.append(t_name[str_loc[0]+1:str_loc[1]])
               elif t_name.count("*") == 2:
                   str_loc = find_all(t_name, "*")
                   tw_names.append(t_name[str_loc[0]+1:str_loc[1]])
               else:
                   tw_names.append(replace_multiple(t_name, rep_list, ""))           
       elif (column[0:8].upper() != "FEEDBACK") and replace_multiple(column.upper(), range(0,10)) not in ["PEER_NAME", "PEER_ID"] and column not in ["Student_name", "Student_id", "ExternalReference"]:
           other_list.append(column)
   tw_list = unique(tw_list)
   tw_names = unique(tw_names)
   
   # Drop redundant rows
   for column in qualtrics:
       del qualtrics[column][0:2] # Drop rows 2 and 3
   # Do I need to drop rows where no student ID exists?  
   
   # Convert the "Finished" column to Boolean values
   for b, case in enumerate(qualtrics["Finished"]):
       qualtrics["Finished"][b] = (case=="True")
       
   # Detect duplicate rows
   for b, case in enumerate(qualtrics["ExternalReference"]):
       if qualtrics["ExternalReference"].count(case) > 1:
           messagebox.showinfo("Error", "Duplicate rows detected. Problem with student ID" + str(case) + ". Please remove duplicate rows and try again.")
           pm.master.destroy()
       else:
           qualtrics["ExternalReference"][b] = cleanupNumber(case) # Convert to an integer
   
   # Convert to a "records" data frame
   qualtrics = df_byRow(qualtrics, "ExternalReference")

   # Collapse the data about peers into a list   
   collapseList(qualtrics, "Peers", "Peer*_id")
   collapseList(qualtrics, "Peer_names", "Peer*_name")

   # Collapse the rating data regarding the teamwork items into a list
   for tw in tw_list: collapseList(qualtrics, tw, tw+"_*")
   collapseList(qualtrics, "TeamWork", tw_list)
   
   # Collapse the feedback data
   collapseList(qualtrics, "Feedback1", "Feedback#1_*_1", delete_blank = False)
   collapseList(qualtrics, "Feedback2", "Feedback#2_*_1", delete_blank = False)
       
   # Eliminate response if no peer ratings were given at all
   delete_list = []
   for student in qualtrics:
       if len(qualtrics[student]["TeamWork"][0]) == 0:
           delete_list.append(student)
   for student in delete_list:
       del qualtrics[student]
          
   # Replace strings with numbers for the TeamWork items; set zero as the bottom score rather than 1
   for student in qualtrics:
       for y, item in enumerate(qualtrics[student]["TeamWork"]):
           for x, rating in enumerate(qualtrics[student]["TeamWork"][y]):
               if type(rating) == str: qualtrics[student]["TeamWork"][y][x] = number_beginning(rating) 
               qualtrics[student]["TeamWork"][y][x] -= 1 # Set zero as the bottom score

   # Upload marks to Canvas
   if config["UploadData"]:
       print("Uploading marks to Canvas...")
       iter = 0
       tstamp = datetime.now().timestamp()
       while (iter==0 and (datetime.now().timestamp() - tstamp < 10)): # Keep looping until at least one iteration has been completed. Timeout after 10 seconds.
           for submission in new_assignment.get_submissions():
               iter = iter + 1
               ID = submission.user_id
               try:
                   print("> " + orig_data[str(ID)]["name"] + ": " + str(round(adj_score(ID),2)))
               except:
                   pass
               if type(adj_score(ID)) is not str:
                   sub_date = submission_date(ID)
                   if (sub_date is not None and ID in qualtrics):
                       submission.edit(submission={'posted_grade': adj_score(ID), 'posted_at': "null", 'submitted_at': datetime.strftime(sub_date, '%Y-%m-%dT23:59:59')})
                   else:
                       submission.edit(submission={'posted_grade': adj_score(ID), 'posted_at': "null"})
               submission.edit(comment={'text_comment': comments(ID)})
       new_assignment.edit(assignment = {"published": config["publish_assignment"]})
       
       if iter==0: 
           messagebox.showinfo("Error", "Unable to upload marks to Canvas. Please try again.")
           pm.master.destroy()

    # Save the data; add all of the penalties information
   if config["ExportData"] and filename != "":
        session["sections"] = GQL.sections()
        print("Saving data in long format...")
        long_data = {}
        long_data["RaterName"] = []
        long_data["ReceiverName"] = []
        long_data["RaterID"] = []
        long_data["ReceiverID"] =  []
        for item in tw_list: long_data[item] = []
        long_data["FeedbackShared"] = []
        long_data["FeedbackInstructor"] = []
        if (config["ScoreType"] == 2):
            long_data["OriginalScore"] = []
            long_data["RatioAdjust"] = []
        long_data["PeerMark"] = []
        long_data["NonComplete"] = []
        for ID2 in orig_data:
            ID = cleanupNumber(ID2)
            if ID in qualtrics:
                long_data["NonComplete"].extend([0] * len(rater_peerID(ID)))
            else:
                long_data["NonComplete"].extend([1] * len(rater_peerID(ID)))
            for a, rater in enumerate(rater_peerID(ID)):
                long_data["ReceiverName"].append(orig_data[str(ID)]["name"])
                long_data["RaterName"].append(orig_data[str(rater)]["name"])
                long_data["ReceiverID"].append(ID)
                long_data["RaterID"].append(rater)
                for b, item in enumerate(tw_list):
                    long_data[item].append(ratings_received(ID)[b][a])
                if a > 0: # Provide the peer feedback
                    long_data["FeedbackShared"].append(feedback(ID)[0][a-1])
                    long_data["FeedbackInstructor"].append(feedback(ID)[1][a-1])
                else: # Provide the person's general feedback
                    if ID in qualtrics:
                        long_data["FeedbackShared"].append(qualtrics[ID]["Feedback1"][-1])
                        long_data["FeedbackInstructor"].append(qualtrics[ID]["Feedback2"][-1])
                    else:
                        long_data["FeedbackShared"].append("")
                        long_data["FeedbackInstructor"].append("")                       
                if (config["ScoreType"] == 2):
                    if "orig_grade" in orig_data[ID2]:
                        long_data["OriginalScore"].append(float(orig_data[ID2]["orig_grade"]))
                    else:
                        long_data["OriginalScore"].append(None)
                    long_data["RatioAdjust"].append(ratee_ratio(ID)[a])
            long_data["PeerMark"].extend([adj_score(ID)] * len(rater_peerID(ID)))

    # Create dataset (wide form) if requested
        print("Saving data in wide format...")
        wide_data = {}
        wide_data["StudentID"] = []
        wide_data["StudentName"] = []
        for section in session["sections"]: wide_data[session["sections"][section]["name"]] = []
        if (config["ScoreType"] == 2):
            wide_data["OriginalScore"] = []
        wide_data["PeerMark"] = []
        wide_data["PartialComplete"] = []
        wide_data["SelfPerfect"] = []
        wide_data["PeersAllZero"] = []
        wide_data["NonComplete"] = []
        for item in other_list: wide_data[item] = []

        for ID2 in orig_data:
            ID = int(ID2)
            if ID in qualtrics:
                wide_data["NonComplete"].append(0)
                for a, item in enumerate(other_list):
                    wide_data[item].append(qualtrics[ID][item])
                wide_data["PartialComplete"].append(PartialComplete(ID))
                wide_data["SelfPerfect"].append(SelfPerfect(ID))
                wide_data["PeersAllZero"].append(PeersAllZero(ID))
            else:
                wide_data["NonComplete"].append(1)
                wide_data["PartialComplete"].append(None)
                wide_data["SelfPerfect"].append(None)
                wide_data["PeersAllZero"].append(None)
                for a, item in enumerate(other_list):
                    wide_data[item].append(None)
            for section in session["sections"]: wide_data[session["sections"][section]["name"]].append(str(section) in orig_data[ID2]["sections"])
            wide_data["StudentName"].append(orig_data[ID2]["name"])
            wide_data["StudentID"].append(ID)
            if (config["ScoreType"] == 2):
                if "orig_grade" in orig_data[ID2]:
                    wide_data["OriginalScore"].append(float(orig_data[ID2]["orig_grade"]))
                else:
                    wide_data["OriginalScore"].append(None)
            wide_data["PeerMark"].append(adj_score(ID))

   if config["ExportData"] and config["UploadData"]:
        messagebox.showinfo("Upload Complete", 'The peer marks have been successfully uploaded to Canvas but are hidden from students. When they are ready to be released to students, you should un-hide them in the Gradebook. Peer rater data saved to "' + filename + '".')
   elif config["UploadData"]:
        messagebox.showinfo("Upload Complete", 'The peer marks have been successfully uploaded to Canvas but are hidden from students. When they are ready to be released to students, you should un-hide them in the Gradebook.')
   elif config["ExportData"]:
        messagebox.showinfo("Export Complete", 'Peer rater data saved to "' + filename + '".')
       
   writeXL(filename = filename, datasets=[long_data, wide_data], sheetnames=["Rater Data", "Student Data"])
   pm.master.destroy()
   
#%% Policies
class Policies():
    def __init__(self, master):
        self.master = master
        self.master.title("Calculate Peer Marks")
        self.master.resizable(0, 0)
        self.master.grab_set()
        
        self.policies = tk.Frame(self.master)
        self.policies.pack()
        self.validation = self.policies.register(only_numbers)
        

                
        labelframe1 = tk.LabelFrame(self.policies, text = "Peer Mark Calculation Policy", fg = "black")
        labelframe1.grid(row = 0, column = 0, pady = 5, padx = 5, sticky = "W")
        
        self.policies_lab1 = tk.Label(labelframe1, text = "Scoring method:", fg = "black").grid(row = 0, column = 0, sticky = "E", pady = 5)
        self.policies_scoring = ttk.Combobox(labelframe1, width = 40, values = ["Average rating received", "Adjusted score based on ratings received"], state="readonly")
        self.policies_scoring.grid(row = 0, column = 1, columnspan = 2, sticky = "W", padx = 5)
        self.policies_scoring.current(config["ScoreType"]-1)
        CreateToolTip(self.policies_scoring, text =
                       '"Average rating received" means each student receives their average mark from their peers.\n' +
                       '"Adjusted score" means the student receives their team mark, which is then adjusted based on their peer ratings.')
        
        tk.Label(labelframe1, text = "Self-scoring policy:", fg = "black").grid(row = 1, column = 0, sticky = "E", pady = 5)
        self.policies_selfrat = ttk.Combobox(labelframe1, width = 40, values = ["Do not count self-rating", "Count self-rating", "Subsitute self-rating with average rating given", "Cap self-rating at average rating given"], state="readonly")
        self.policies_selfrat.grid(row = 1, column = 1, columnspan = 2, sticky = "W", padx = 5)
        self.policies_selfrat.current(config["SelfVote"])
        CreateToolTip(self.policies_selfrat, text =
                       'This option is used to specify whether to count each student\'s self-rating\n'+
                       'as part of their peer mark.')
        
        tk.Label(labelframe1, text = "Minimum peer responses:", fg = "black").grid(row = 2, column = 0, sticky = "E", pady = 5)
        self.policies_minimum = tk.Entry(labelframe1, width = 5, validate="key", validatecommand=(self.validation, '%S'))
        self.policies_minimum.grid(row = 2, column = 1, padx = 5, sticky = "W")
        self.policies_minimum.insert(0, str(config["MinimumPeers"]))
        CreateToolTip(self.policies_minimum, text =
                       'Specifies the minimum number of peers required to calculate the peer rating.\n' +
                       'Under the "Adjusted score" method, the student receives their unadjusted team mark.')
        
        labelframe2 = tk.LabelFrame(self.policies, text = "Penalties Policy", fg = "black")
        labelframe2.grid(row = 1, column = 0, pady = 5, padx = 5, sticky = "W")
        tk.Label(labelframe2, text = "Apply % penalty", fg = "black").grid(row = 0, column = 2, sticky = "W")
        
        tk.Label(labelframe2, text = "Students do not complete the peer evaluation", fg = "black").grid(row = 1, rowspan = 2, column = 0, sticky = "NE")
        tk.Label(labelframe2, text = "\n", fg = "black").grid(row = 1, rowspan = 2, column = 1, sticky = "NE")
        self.penalty_noncomplete = tk.Entry(labelframe2, width = 5, validate="key", validatecommand=(self.validation, '%S'))
        self.penalty_noncomplete.grid(row = 1, column = 2, sticky="W", padx = 5, pady = 5)
        self.penalty_noncomplete.insert(0, config["Penalty_NonComplete"])
        CreateToolTip(self.penalty_noncomplete, text =
                       'Automatically deduct % if the student does not answer any questions in the peer evaluation.')

        tk.Label(labelframe2, text = "Students only partially complete the peer evaluation", fg = "black").grid(row = 2, rowspan = 2, column = 0, sticky = "NE")
        tk.Label(labelframe2, text = "\n", fg = "black").grid(row = 2, rowspan = 2, column = 1, sticky = "NE")
        self.penalty_partialcomplete = tk.Entry(labelframe2, width = 5, validate="key", validatecommand=(self.validation, '%S'))
        self.penalty_partialcomplete.grid(row = 2, column = 2, sticky="W", padx = 5, pady = 5)
        self.penalty_partialcomplete.insert(0, config["Penalty_PartialComplete"])
        CreateToolTip(self.penalty_partialcomplete, text =
                       'Automatically deduct % if the student completes only some of the peer evaluation questions but does not finish the survey.')
        
        tk.Label(labelframe2, text = "Students submit the peer evaluation late", fg = "black").grid(row = 3, rowspan = 2, column = 0, sticky = "NE")
        tk.Label(labelframe2, text = "(per day)", fg = "black").grid(row = 3, column = 2, sticky = "E")
        self.penalty_late = tk.Entry(labelframe2, width = 5, validate="key", validatecommand=(self.validation, '%S'))
        self.penalty_late.grid(row = 3, column = 2, sticky="W", padx = 5, pady = 5)
        self.penalty_late.insert(0, config["Penalty_PerDayLate"])
        CreateToolTip(self.penalty_late, text =
                       'Automatically deduct % for each day the peer evaluation is submitted after the due date.')
        
        tk.Label(labelframe2, text = "Students give themselves perfect scores", fg = "black").grid(row = 4, rowspan = 2, column = 0, sticky = "NE")
        self.perfect_score = tk.IntVar()
        ttk.Radiobutton(labelframe2, text = "Retain scores", variable = self.perfect_score, value = 0).grid(row = 4, column = 1, sticky = "W")
        ttk.Radiobutton(labelframe2, text = "Exclude scores", variable = self.perfect_score, value = 1).grid(row = 5, column = 1, sticky = "W")
        self.perfect_score.set(config["Exclude_SelfPerfect"])
        self.penalty_selfperfect = tk.Entry(labelframe2, width = 5, validate="key", validatecommand=(self.validation, '%S'))
        self.penalty_selfperfect.grid(row = 4, column = 2, sticky="W", padx = 5)
        self.penalty_selfperfect.insert(0, config["Penalty_SelfPerfect"])
        CreateToolTip(self.penalty_selfperfect, text =
                       'Automatically deduct % if the student gives themselves a perfect self-rating across all questions.')
        
        tk.Label(labelframe2, text = "Students give all peers the bottom score", fg = "black").grid(row = 6, rowspan = 2, column = 0, sticky = "NE")
        self.bottom_score = tk.IntVar()
        ttk.Radiobutton(labelframe2, text = "Retain scores", variable = self.bottom_score, value = 0).grid(row = 6, column = 1, sticky = "W")
        ttk.Radiobutton(labelframe2, text = "Exclude scores", variable = self.bottom_score, value = 1).grid(row = 7, column = 1, sticky = "W")
        self.bottom_score.set(config["Exclude_PeersAllZero"])
        self.penalty_allzero = tk.Entry(labelframe2, width = 5, validate="key", validatecommand=(self.validation, '%S'))
        self.penalty_allzero.grid(row = 6, column = 2, sticky="W", padx = 5)
        self.penalty_allzero.insert(0, config["Penalty_PeersAllZero"])
        CreateToolTip(self.penalty_allzero, text =
                       'Automatically deduct % if the student gives all of their peers the lowest rating across all questions.')        
        
        labelframe3 = tk.LabelFrame(self.policies, text = "Additional Settings", fg = "black")
        labelframe3.grid(row = 2, column = 0, pady = 5, padx = 5, sticky = "W")

        tk.Label(labelframe3, text = "Peer mark scale ranges from 1 to:", fg = "black").grid(row = 0, column = 0, sticky = "NE")
        self.points_possible = tk.Entry(labelframe3, width = 5, validate="key", validatecommand=(self.validation, '%S'))
        self.points_possible.grid(row = 0, column = 1, sticky="W", padx = 5)
        self.points_possible.insert(0, config["PointsPossible"])
        CreateToolTip(self.points_possible, text =
                       'Only change this setting if you modify the scale of the peer evaluation survey questions.\n'+
                       'For example: you change the questions from a 1 to 5 scale to a 1 to 7 scale.')

        tk.Label(labelframe3, text = "Rescale peer mark to score out of:", fg = "black").grid(row = 1, column = 0, sticky = "NE")
        self.rescale_to = tk.Entry(labelframe3, width = 5, validate="key", validatecommand=(self.validation, '%S'))
        self.rescale_to.grid(row = 1, column = 1, sticky="W", padx = 5)
        self.rescale_to.insert(0, config["RescaleTo"])
        CreateToolTip(self.rescale_to, text =
                       'Use this setting to rescale the Peer Mark. For example, if you want the Peer Mark to\n'+
                       'be a score of 10, write \'10\' in the box.')                      
        
        tk.Label(labelframe3, text = "Assignment name:", fg = "black").grid(row = 2, column = 0, sticky = "E", pady = 5)
        self.pm_name = tk.Entry(labelframe3, width = 40)
        self.pm_name.grid(row = 2, column = 1, padx = 5, sticky = "W")
        self.pm_name.insert(0, str(config["PeerMark_Name"]))
        CreateToolTip(self.pm_name, text =
                       'The name that the Peer Mark will be given in the Canvas grade book.')

        self.publish_assignment = tk.IntVar()
        self.publish_assignment.set(int(config["publish_assignment"]))
        tk.Label(labelframe3, text = "Publish Peer Mark column in Grade Book?", fg = "black").grid(row = 3, column = 0, sticky = "NE")
        self.pub_assn = ttk.Checkbutton(labelframe3, text = "Make the column visible but hide grades", variable = self.publish_assignment)
        self.pub_assn.grid(row = 3, column = 1, sticky = "W")
        CreateToolTip(self.pub_assn, text =
                       'Whether the Peer Evaluation will be published in Canvas.')

        self.weekdays_only = tk.IntVar()
        self.weekdays_only.set(int(config["weekdays_only"]))
        tk.Label(labelframe3, text = "Exclude weekends?", fg = "black").grid(row = 4, column = 0, sticky = "NE")
        self.wko = ttk.Checkbutton(labelframe3, text = "Only count weekdays when calculating late penalty", variable = self.publish_assignment)
        self.wko.grid(row = 4, column = 1, sticky = "W")
        CreateToolTip(self.wko, text =
                       'When calculating penalties for late submission of the Peer Evaluation, only count weekdays.')

        self.export_data = tk.IntVar()
        self.export_data.set(int(config["ExportData"]))
        tk.Label(labelframe3, text = "Export peer data?", fg = "black").grid(row = 5, column = 0, sticky = "NE")
        self.export = ttk.Checkbutton(labelframe3, text = "Export peer ratings to an XLSX file", variable = self.export_data)
        self.export.grid(row = 5, column = 1, sticky = "W")
        CreateToolTip(self.export, text =
                       'After the peer marks are calculated, export them to an Excel file.\n' +
                       'This format is useful for inspecting the data and performing data analyses.')
        
        self.upload_data = tk.IntVar()
        self.upload_data.set(int(config["UploadData"]))
        tk.Label(labelframe3, text = "Upload peer data?", fg = "black").grid(row = 6, column = 0, sticky = "NE")
        self.upload = ttk.Checkbutton(labelframe3, text = "Upload peer marks and feedback to Canvas", variable = self.upload_data)
        self.upload.grid(row = 6, column = 1, sticky = "W")
        CreateToolTip(self.wko, text =
                       'Upload the peer marks and feedback to the Grade book in Canvas.\n' +
                       'This creates a new data column (it does not overwrite existing grades).')
        
        self.save_policies = tk.IntVar()
        self.save_policies.set(1)
        tk.Label(labelframe3, text = "Retain policies?", fg = "black").grid(row = 7, column = 0, sticky = "NE")
        ttk.Checkbutton(labelframe3, text = "Save policies for future assessments", variable = self.save_policies).grid(row = 7, column = 1, sticky = "W")
        
        tk.Button(self.policies, text = "Continue", fg = "black", command = self.save_settings).grid(row = 13, column = 0, columnspan = 3, pady = 5)
        
        self.policies_scoring.bind("<<ComboboxSelected>>", self.change_assnname)
    
    def change_assnname(self, event_list = ""):
        self.pm_name.delete(0, 'end')
        if self.policies_scoring.current()==0:
            self.pm_name.insert(0, "Peer Mark (average mark received)") # Average mark received
        else:
            self.pm_name.insert(0, "Team Assignment (peer-rating adjusted)")
    
    def save_settings(self):
        config["ScoreType"] = int(self.policies_scoring.current()+1)
        config["SelfVote"] = int(self.policies_selfrat.current())
        config["MinimumPeers"] = int(self.policies_minimum.get())
        config["PeerMark_Name"] = str(self.pm_name.get().strip())
        config["Penalty_NonComplete"] = int(self.penalty_noncomplete.get())
        config["Penalty_PartialComplete"] = int(self.penalty_partialcomplete.get())
        config["Penalty_PerDayLate"] = int(self.penalty_late.get())
        config["Penalty_SelfPerfect"] = int(self.penalty_selfperfect.get())
        config["Penalty_PeersAllZero"] = int(self.penalty_allzero.get())
        config["PointsPossible"] = int(self.points_possible.get())
        config["RescaleTo"] = int(self.rescale_to.get())
        config["publish_assignment"] = bool(self.publish_assignment.get())
        config["weekdays_only"] = bool(self.weekdays_only.get())
        config["ExportData"] = bool(self.export_data.get())
        config["UploadData"] = bool(self.upload_data.get())
        
        if self.save_policies.get() == 1: save_config()
        self.master.destroy()
        
def Policies_call():
    policies = tk.Toplevel()
    pol = Policies(policies)

#%% Bulk mail
class BulkMail:
    def __init__(self, master, preset = "", subject = "", message = ""):
        self.preset = preset
        self.master = master
        self.master.title("Send Bulk Message via Canvas")
        self.master.resizable(0, 0)
        
        self.master.bind("<FocusIn>", self.check_focus)
        
        self.window = tk.Frame(self.master)
        self.window.pack()
        
        tk.Label(self.window, text = "Unit:", fg = "black").grid(row = 1, column = 0, sticky = "E", pady = 5)
        self.window_unit = ttk.Combobox(self.window, width = 60, state="readonly")
        self.window_unit.grid(row = 1, column = 1, sticky = "W", padx = 5)
        self.window_unit["values"] = tuple(session["course.names"])
        self.window_unit.current(0)

        self.sec_text = tk.StringVar()
        self.sec_text.set("Distribution list file")
        self.secondary_lab = tk.Label(self.window, textvariable = self.sec_text, fg = "black").grid(row = 3, column = 0, sticky = "E", pady = 5)        

        # Distribution list file
        self.window_distlist = tk.Button(self.window, text = "Select...", fg = "black", command = self.select_file)
        self.window_distlist.grid(row = 3, column = 1, pady = 5, padx = 5, sticky = "W")
        self.datafile = tk.Label(self.window, text = "", fg = "black")
        self.datafile.grid(row = 4, column = 1, pady = 5, sticky = "W")

        # Group sets
        self.window_groupset = ttk.Combobox(self.window, width = 60, state="readonly")
        self.group_sets(self)
        self.window_unit.bind("<<ComboboxSelected>>", self.group_sets)
        #self.groupset.grid(row = 3, column = 1, sticky = "W", padx = 5)

        tk.Label(self.window, text = "Send to (via):", fg = "black").grid(row = 0, column = 0, sticky = "E", pady = 5)
        self.window_sendto = ttk.Combobox(self.window, width = 60, state="disabled")
        self.window_sendto.grid(row = 0, column = 1, sticky = "W", pady = 5, padx = 5)
        self.window_sendto["values"] = ["All students (Canvas)",
                                        "Students in groups (Canvas)",
                                        "Students not in groups (Canvas)",
                                        "Groups (Canvas)",
                                        "All students (distribution list file)",
                                        "Students who completed the survey (distribution list file)",
                                        "Students who did not complete the survey (distribution list file)"]
        self.change_message(self)
        self.window_sendto.bind("<<ComboboxSelected>>", self.change_message)
        
        tk.Label(self.window, text = "Message:", fg = "black").grid(row = 5, column = 0, columnspan = 2, pady = 5, padx = 5)

        tk.Label(self.window, text = "Subject:", fg = "black").grid(row = 6, column = 0, sticky = "E", pady = 5)
        self.subject_text = tk.StringVar()
        self.subject = tk.Entry(self.window, textvariable = self.subject_text)
        self.subject.grid(row = 6, column = 1, pady = 5, padx = 5, sticky = "WE")
        self.subject_text.set(subject)

        self.message = tk.Text(self.window, width = 60, height = 15, wrap = tk.WORD)
        self.message.grid(row = 7, column = 0, columnspan = 2, sticky = "W", pady = 5, padx = 5)
        self.message.bind("<KeyRelease>", self.check_status)
        self.message.insert(tk.END, message)
        
        self.button_field = tk.Button(self.window, text = "Fields...", fg = "black", command = self.field_picker)
        self.button_field.grid(row = 8, column = 0, columnspan = 2, sticky = "W", padx = 5)
        
        self.sendmessage = tk.Button(self.window, text = "Send Bulk Message", fg = "black", state = "disabled", command = self.MessageSend)
        self.sendmessage.grid(row = 8, column = 0, columnspan = 2, pady = 5)
       
        self.distlist = {}
        session["DistList"] = ""

        self.save_template = tk.IntVar()
        self.st_check = ttk.Checkbutton(self.window, text = "Save message as template", variable = self.save_template, state = "normal")
        self.st_check.grid(row = 8, column = 0, columnspan = 2, sticky = "E")
        
        # Check presets
        if preset=="confirm":
            self.window_sendto.current(3)
        elif preset=="nogroup":
            self.window_sendto.current(2)
        elif preset=="invitation":
            self.window_sendto.current(4)
        elif preset=="reminder":
            self.window_sendto.current(6)
        else:
            self.window_sendto.current(0)
            self.window_sendto["state"] = "readonly"
            self.st_check["state"] = "disabled"

        self.change_message(self)
        self.check_status(self)
    
    def check_focus(self, event_type = ""):
        self.master.grab_set()
    
    def distlist_keys(self, event_type = ""):
        keys = []
        if self.window_sendto.current() in [0,1,2]:
            keys = ["first_name", "last_name", "email"]
        if self.window_sendto.current() == 1:
            keys.append("group_name")
            keys.append("peers_list")
            keys.append("peers_bulletlist")
            keys.append("group_homepage")
        elif self.window_sendto.current() == 3:
            keys = ["firstnames_list", "members_list", "members_bulletlist", "name", "group_homepage"]        
        elif self.window_sendto.current() in [4,5,6]:
            self.create_distlist()
            if len(self.distlist) > 0:
                keys = list(self.distlist[next(iter(self.distlist))].keys())
        return(keys)
    
    def field_picker(self, event_type = ""):
        self.master.grab_release()
        field_window = tk.Toplevel()
        fp = FieldPicker(field_window, self.distlist_keys())
    
    def insert_field(self, event_type = "", field = ""):
        self.message.insert(tk.INSERT, field)
    
    def check_status(self, event_type = ""):
        if (len(self.message.get(1.0, tk.END).strip()) < 1) or (self.window_sendto.current() in [4,5,6] and session["DistList"] == ""):
            self.sendmessage["state"] = "disabled"
        else:
            self.sendmessage["state"] = "normal"
        if (self.window_sendto.current() in [4,5,6] and session["DistList"] == ""):
            self.button_field["state"] = "disabled"
        else:
            self.button_field["state"] = "normal"
        
    def select_file(self):
       session["DistList"] = filedialog.askopenfilename(initialdir = ".", title = "Select file",filetypes = (("Compatible File Types",("*.csv")),("All Files","*.*")))
       self.datafile["text"] = session["DistList"]
       self.check_status(self)
       
    def change_message(self, event_type = ""):
        self.distlist = {}
        if self.window_sendto.current() == 0: #Canvas all students
            self.sec_text.set("")
            self.window_distlist.grid_forget()
            self.window_groupset.grid_forget()
        elif self.window_sendto.current() in [1,2,3]:
            self.sec_text.set("Group set:")
            self.window_distlist.grid_forget()
            self.window_groupset.grid(row = 3, column = 1, sticky = "W", padx = 5)
        elif self.window_sendto.current() in [4,5,6]:
            self.sec_text.set("Distribution list file:")
            self.window_distlist.grid(row = 3, column = 1, pady = 5, padx = 5, sticky = "W")
            self.window_groupset.grid_forget()
            
    def group_sets(self, event_type = ""):
        self.group_sets_id = []
        self.group_sets_name = []
        GQL.course_id = session["course.ids"][self.window_unit.current()]
        for group_set in GQL.group_sets():
            self.group_sets_id.append(group_set["_id"])
            self.group_sets_name.append(group_set["name"])
        if len(self.group_sets_id) > 0:
            self.window_groupset["values"] = tuple(self.group_sets_name)
            self.window_groupset.current(0)
        else:
            self.window_groupset["values"] = ()
            self.window_groupset.set('')            
        session["group_set"] = self.group_sets_id

    def create_distlist(self, event_type = ""):
        GQL.course_id = session["course.ids"][self.window_unit.current()] # Set the course ID to the one currently chosen
        if self.window_sendto.current() in [1,2,3]: # Prepopulate the list of groups
            GQL.groupset_id = self.group_sets_id[self.window_groupset.current()]        
        if self.window_sendto.current() in [0,1,2]: # Download students from Canvas
            students = GQL.students_comprehensive()
        if self.window_sendto.current() == 0: # Simply copy the distribution list
            self.distlist = students
        if self.window_sendto.current() == 1: # Drop students without groups; get list of peers
            for student in students:
                if "group_id" in students[student]:
                    self.distlist[student] = students[student]
                    self.distlist[student]["peers_bulletlist"] = ""
                    self.distlist[student]["peers_list"] = ""
                    for n, peer in enumerate(self.distlist[student]["peers"], 1):
                        self.distlist[student]["peers_bulletlist"] += "* " + students[peer]["name"]
                        if (n < len(self.distlist[student]["peers"])):
                            self.distlist[student]["peers_bulletlist"] += "\n"
                            if (n > 1):
                                self.distlist[student]["peers_list"] += ", "
                        elif (n == len(self.distlist[student]["peers"])):
                            self.distlist[student]["peers_list"] += " and "
                        self.distlist[student]["peers_list"] += students[peer]["name"]
                    self.distlist[student]["group_homepage"] = config["API_URL"] + "groups/" + str(students[student]["group_id"]) + "/"
        if self.window_sendto.current() == 2: # Drop students who have groups
            for student in students:
                if "group_id" not in students[student]:
                    self.distlist[student] = students[student]
        if self.window_sendto.current() == 3: # Get list of groups
            groups = GQL.groups()
            students = GQL.students_comprehensive()
            for group in groups:
                if len(groups[group]["users"]) > 0:
                    self.distlist[group] = groups[group]
                    self.distlist[group]["members_bulletlist"] = ""
                    self.distlist[group]["members_list"] = ""
                    self.distlist[group]["firstnames_list"] = ""
                    self.distlist[group]["group_homepage"] = config["API_URL"] + "groups/" + str(group) + "/"
                    for n, member in enumerate(groups[group]["users"],1):
                        self.distlist[group]["members_bulletlist"] += "* " + students[member]["name"]
                        if (n < len(groups[group]["users"])):
                            self.distlist[group]["members_bulletlist"] += "\n"
                            if (n > 1):
                                self.distlist[group]["members_list"] += ", "
                                self.distlist[group]["firstnames_list"] += ", "
                        elif (n == len(groups[group]["users"])):
                            self.distlist[group]["members_list"] += " and "
                            self.distlist[group]["firstnames_list"] += " and "
                        self.distlist[group]["members_list"] += students[member]["name"]
                        self.distlist[group]["firstnames_list"] += students[member]["first_name"]
        if self.window_sendto.current() in [4,5,6]:
            print(session["DistList"])
            if len(session["DistList"]) > 0:
                try:
                    df = readTable(filename = session["DistList"])
                    students = {}
                    session["ID_column"] = None
                    if "Student_id" in df: session["ID_column"] = "Student_id"
                    if "External Data Reference" in df: session["ID_column"] = "External Data Reference"
                    if "ExternalDataReference" in df: session["ID_column"] = "ExternalDataReference"
                    if "id" in df: session["ID_column"] = "id"
                    if "ID" in df: session["ID_column"] = "ID"
                    if session["ID_column"] is not None:
                        df = df_byRow(df, id_col = session["ID_column"])
                        for ID in df:
                            students[str(ID)] = {}
                            for col_name in df[ID]:
                                students[str(ID)][col_name] = df[ID][col_name]                        
                    else:
                        messagebox.showinfo("Error", "Cannot find user ID column in file. This column should be labelled 'id', 'Student_id' or 'External Data Reference'.")
                        self.master.destroy()
                except:
                    messagebox.showinfo("Error", "Trouble parsing file.")
                    self.master.destroy()
        if self.window_sendto.current() == 4:
            self.distlist = students
        if self.window_sendto.current() in [5,6] and "Status" not in df[list(df.keys())[0]]: # Look for "Status" in the first dictionary key
            messagebox.showinfo("Error", "Cannot find 'Status' column in file used to indicate whether a student has completed the survey.")
            self.master.destroy()
        if self.window_sendto.current() == 5:
            for student in students:
                if students[student]["Status"] == "Finished Survey":
                    self.distlist[student] = students[student]
        if self.window_sendto.current() == 6:
            for student in students:
                if students[student]["Status"] != "Finished Survey":
                    self.distlist[student] = students[student]
        
        # Delete all of the fields that are lists
        delete_dict = {}
        for case in self.distlist:
            for field in self.distlist[case]:
                if type(self.distlist[case][field]) in [list,tuple,dict]:
                    if case not in delete_dict: delete_dict[case] = []
                    delete_dict[case].append(field)
        for case in delete_dict:
            for field in delete_dict[case]:
                del self.distlist[case][field]
                
    def MessageSend(self):
        if self.preset!="" and self.save_template.get() == 1:
            print("Saving message as template...")
            config["subject_" + self.preset] = self.subject_text.get()
            config["message_" + self.preset] = self.message.get(1.0, tk.END)
            save_config()
            
        print("Preparing distribution list...")                
        self.create_distlist(self)
        if len(self.distlist) == 0: 
            messagebox.showinfo("Error", "No students found.")
            self.master.destroy()
        else:
            print("Preparing messages...")
            messagelist = {}
            for case in self.distlist:
                messagelist[case] = {}
                message = self.message.get(1.0, tk.END)
                subject = self.subject_text.get()
                for field in self.distlist[case]:
                    subject = subject.replace("["+field+"]", str(self.distlist[case][field]))
                    message = message.replace("["+field+"]", str(self.distlist[case][field]))
                messagelist[case]["subject"] = subject
                messagelist[case]["message"] = message
            
            for k, case in enumerate(messagelist, 1):
                print("Sending message {} of {}".format(k, len(messagelist)))
                if self.window_sendto.current() == 3: # Send group message
                    bulk_message = canvas.create_conversation(recipients = ['group_' + str(case)],
                                                         body = messagelist[case]["message"],
                                                         subject = messagelist[case]["subject"],
                                                         group_conversation = True,
                                                         context_code = "course_" + str(GQL.course_id))            
                else: # Send individual message
                    try: # Try sending the course code information
                        bulk_message = canvas.create_conversation(recipients = [str(case)],
                                                                  body = messagelist[case]["message"],
                                                                  subject = messagelist[case]["subject"],
                                                                  force_new = True,
                                                                  context_code = "course_" + str(GQL.course_id))
                    except: # Ignore the course code; don't send within context
                        bulk_message = canvas.create_conversation(recipients = [str(case)],
                                                                  body = messagelist[case]["message"],
                                                                  subject = messagelist[case]["subject"],
                                                                  force_new = True)
            messagebox.showinfo("Finished", "All messages have been sent.")
        self.master.destroy()

def BulkMail_call(preset = ""):
    global bm
    window = tk.Toplevel()
    bm = BulkMail(window)
    
def BulkMail_confirm():
    global bm
    window = tk.Toplevel()
    bm = BulkMail(window,
                  preset = "confirm", 
                  subject = config["subject_confirm"],
                  message = config["message_confirm"])

def BulkMail_nogroup():
    global bm
    window = tk.Toplevel()
    bm = BulkMail(window,
                  preset = "nogroup", 
                  subject = config["subject_nogroup"],
                  message = config["message_nogroup"])
    
def BulkMail_invitation():
    global bm
    window = tk.Toplevel()
    bm = BulkMail(window,
                  preset = "invitation", 
                  subject = config["subject_invitation"],
                  message = config["message_invitation"])

def BulkMail_reminder():
    global bm
    window = tk.Toplevel()
    bm = BulkMail(window,
                  preset = "reminder", 
                  subject = config["subject_reminder"],
                  message = config["message_reminder"])

#%% Field picker
class FieldPicker:
    def __init__(self, master, fields = []):
        self.fields = fields
        self.master = master
        self.master.title("Fields")
        self.master.resizable(0,0)
        self.master.grab_set()
        
        self.window = tk.Frame(self.master)
        self.window.pack()
        
        self.field_list = tk.Listbox(self.window, width = 30, height = 20)
        self.field_list.pack(padx = 10, pady = 10)
        for field in fields:
            self.field_list.insert(tk.END, field)
            
        self.insert_field = tk.Button(self.window, text = "Insert Field", command = self.insert_text)
        self.insert_field.pack(padx = 10, pady = 10)
        
    def insert_text(self):
        self.master.grab_release()
        bm.insert_field(field = "[" + self.fields[self.field_list.curselection()[0]] + "]")
        self.master.destroy()  

#%% Qualtrics Template
def upload_template():
    subprocess.Popen(['explorer', os.path.abspath(os.path.dirname(sys.argv[0])) + "\\survey"])
    messagebox.showinfo("Create Survey", "To create the survey, upload the \"Peer_Evaluation_Qualtrics_Template.qsf\" template to Qualtrics. Read the \"Instructions.txt\" file for more information.")

#%% Main Menu

class MainMenu:
    def __init__(self, master):
        self.master = master
        self.master.title("PEER")
        #self.window.geometry("300x250") # size of the window width:- 500, height:- 375
        self.master.resizable(0, 0) # this prevents from resizing the window
        self.master.bind("<FocusIn>", self.check_status)

        self.window = tk.Frame(self.master)
        self.window.pack(side="top", pady = 0, padx = 50)

        # create a pulldown menu, and add it to the menu bar
        self.menubar = tk.Menu(self.window)
        self.master.config(menu = self.menubar)

        self.filemenu = tk.Menu(self.menubar, tearoff=0)
        self.filemenu.add_command(label="Settings", command = Settings_call)
        self.filemenu.add_separator()
        self.filemenu.add_command(label="Exit", command = self.exit_program)
        self.menubar.add_cascade(label="File", menu=self.filemenu)              

        self.toolmenu = tk.Menu(self.menubar, tearoff = 0)
        self.toolmenu.add_command(label="Send Bulk Message", command = BulkMail_call)
        self.toolmenu.add_command(label="Download Class List", command = ExportGroups_classlist)
        self.menubar.add_cascade(label="Tools", menu=self.toolmenu)

        labelframe1 = tk.LabelFrame(self.window, text = "Group Formation", fg = "black")
        labelframe1.pack(pady = 10)

        self.button_cgm = tk.Button(labelframe1, text = "Confirm group\nmembership", fg = "black", command = BulkMail_confirm, width = 20)
        self.button_cgm.pack(pady = 10, padx = 10)# 'text' is used to write the text on the Button
        CreateToolTip(self.button_cgm, text = 
                      'Contact students via the Canvas bulk mailer to confirm\n'+
                      'their membership in each group, as well as their peers.')

        self.button_ng = tk.Button(labelframe1, text = "Contact students\nwithout groups", fg = "black", command = BulkMail_nogroup, width = 20)
        self.button_ng.pack(pady = 10, padx = 10)# 'text' is used to write the text on the Button
        CreateToolTip(self.button_ng, text = 
                      'Contact students via the Canvas bulk mailer\n'+
                      'who are not currently enrolled in groups.')

        labelframe2 = tk.LabelFrame(self.window, text = "Survey Launch", fg = "black")
        labelframe2.pack(pady = 10)
        
        self.button_tp = tk.Button(labelframe2, text = "Create Qualtrics survey\nfrom template", fg = "black", command = upload_template, width = 20)
        self.button_tp.pack(pady = 10, padx = 10)# 'text' is used to write the text on the Button
        CreateToolTip(self.button_tp, text = 
                      'Upload the Peer Evaluation template\n'+
                      'to Qualtrics to create the survey.')
        
        self.button_dl = tk.Button(labelframe2, text = "Create Contacts list\nfrom group membership", fg = "black", command = ExportGroups_call, width = 20)
        self.button_dl.pack(pady = 10, padx = 10)# 'text' is used to write the text on the Button
        CreateToolTip(self.button_dl, text = 
                      'Export group membership data from Canvas and create\n'+
                      'a Contacts list file to be uploaded to Qualtrics.')

        self.button_si = tk.Button(labelframe2, text = "Send survey invitations\nfrom distribution list", fg = "black", command = BulkMail_invitation, width = 20)
        self.button_si.pack(pady = 10, padx = 10)# 'text' is used to write the text on the Button
        CreateToolTip(self.button_si, text = 
                      'Send students a unique link to the Peer Evaluation from a\n'+ 
                      'Qualtrics distribution list using the Canvas bulk mailer.')


        labelframe3 = tk.LabelFrame(self.window, text = "Peer Evaluation Finalisation", fg = "black")
        labelframe3.pack(pady = 10)        
    
        self.button_sr = tk.Button(labelframe3, text = "Send survey reminders\nfrom distribution list", fg = "black", command = BulkMail_reminder, width = 20)
        self.button_sr.pack(pady = 10, padx = 10)# 'text' is used to write the text on the Button
        CreateToolTip(self.button_sr, text = 
                      'Send students who haven\'t completed the Peer Evaluation a reminder to\n'+
                      'complete it (from a Qualtrics distribution list via the Canvas bulk mailer).')

        self.button_pm = tk.Button(labelframe3, text = "Upload peer marks\nand comments", fg = "black", command = PeerMark_call, width = 20)
        self.button_pm.pack(pady = 10)# 'text' is used to write the text on the Button     
        CreateToolTip(self.button_pm, text = 
                      'Calculate peer marks from the Qualtrics survey data and upload\n'+ 
                      'these marks and peer feedback to the Canvas grade book.')

        self.check_status()
        
        if os.path.abspath(os.path.dirname(sys.argv[0]))[-4:].lower() == ".zip":
            messagebox.showinfo("Error", "PEER cannot be run within a ZIP folder. Please extract all of the files from the ZIP and try again.")
            mm.destroy()
            sys.exit()
        
        if config["API_TOKEN"] == "" or config["API_URL"] == "":
            messagebox.showinfo("Warning", "No configuration found. Please go to File -> Settings to set up a link to Canvas.")            
    
    def check_status(self, event_type = ""):
        global canvas, GQL
        if config["API_TOKEN"] == "" or config["API_URL"] == "":
            self.button_dl["state"] ='disabled'
            self.button_pm["state"] ='disabled'
            self.button_ng["state"] = 'disabled'
            self.button_cgm["state"] = 'disabled'
            self.button_si["state"] = 'disabled'
            self.button_sr["state"] = 'disabled'
            self.button_pm["state"] = 'disabled'
            self.toolmenu.entryconfigure("Send Bulk Message", state = 'disabled')
            self.toolmenu.entryconfigure("Download Class List", state = 'disabled')
        else:
            GQL = GraphQL()
            if "course.names" not in session:
                session["course.names"] = []
                session["course.ids"] = []
                courses = GQL.courses()
                for course in courses:
                    session["course.names"].append(courses[course]["name"])
                    session["course.ids"].append(course)
            canvas = Canvas(cleanupURL(config["API_URL"]), config["API_TOKEN"])
            self.button_dl["state"] ='normal'
            self.button_pm["state"] ='normal'
            self.button_ng["state"] = 'normal'
            self.button_cgm["state"] = 'normal'
            self.button_si["state"] = 'normal'
            self.button_sr["state"] = 'normal'
            self.button_pm["state"] = 'normal'
            self.toolmenu.entryconfigure("Send Bulk Message", state = 'normal')
            self.toolmenu.entryconfigure("Download Class List", state = 'normal')

            
    def exit_program(self):
        self.master.destroy()
            
temp_data = {}
canvas = None
bm = None
pm = None

long_data = {}
wide_data = {}

top_frame = tk.Tk()
mm = MainMenu(top_frame)
top_frame.mainloop()



